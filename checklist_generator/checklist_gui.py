import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import threading
from pathlib import Path
from datetime import datetime
PENDING_HEADERS = [
    "Model", "Product Desc", "Country of Origin",
    "CTH", "Basic Duty Notn", "Basic Duty Notn SNo", "Basic Duty Rate", 
    "Customs Health Cess Notn", "Customs Health Cess SNo", "Customs Health Cess Rate", 
    "SWS Notification", "SWS Notification SrNo", "SWS Rate", "IGST Notification", 
    "IGST Notification SrNo", "IGST Rate", "AIDC Notn (Customs)", "AIDC Notn Sr.No.(Customs)", 
    "End Use", "Generic Description", "Added_By", "Added_At", "Importer",
    "Status", "Reviewed_By", "Reviewed_At", "Reviewer_Remarks"
]

MASTER_SHEET_COLUMNS = [
    "Job No", "BE Date", "Model", "Product Desc", "CTH", 
    "Basic Duty Notn", "Basic Duty Notn SNo", "Basic Duty Rate", 
    "Customs Health Cess Notn", "Customs Health Cess SNo", "Customs Health Cess Rate", 
    "SWS Notification", "SWS Notification SrNo", "SWS Rate", 
    "IGST Notification", "IGST Notification SrNo", "IGST Rate", 
    "AIDC Notn (Customs)", "AIDC Notn Sr.No.(Customs)", 
    "End Use", "Generic Description", "Country of Origin"
]

# Import tksheet
from tksheet import Sheet

# Adjust path to import auth and logic
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import auth
import checklist_logic

# Brand Constants
_PRIMARY_BLUE = "#1F3F6E"; _ACCENT_RED = "#D8232A"; _DARK_TEXT = "#1F2937"
_MUTED_GRAY = "#6B7280"; _LIGHT_BG = "#F4F6F8"; _PANEL_WHITE = "#FFFFFF"
_BORDER_GRAY = "#E5E7EB"; _HOVER_BLUE = "#2A528F"; _HEADER_BG = "#D6E4F0"

# 1-based indices of editable manual columns
EDITABLE_INDICES = {1, 2, 3, 5, 6, 7, 8, 16, 23, 32}
# 1-based indices of master-sourced columns
MASTER_INDICES = {4, 9, 10, 11, 22, 24, 25, 41, 42, 54, 55, 81, 82}

# Column Widths classifications
NARROW_COLS = {'Quantity', 'Unit', 'Currency', 'Rate', 'CTH', 'Sr no', 'IGST SrNo', 'SWS SrNo'}
MEDIUM_COLS = {'Inv No', 'Date', 'TOI', 'Model', 'BRAND', 'Basic notification', 'IGST Notn No', 'SWS Notn No'}
WIDE_COLS = {'Product Desc', 'Generic Description', 'END USE', 'Country of origin'}

def wrap_header(text, max_len=15):
    """Returns the cleaned header text on a single line (no wrapping)."""
    return str(text).strip()

# Pre-calculate column widths
GRID_WIDTHS = []
for h in checklist_logic.LOGISYS_HEADERS:
    h_clean = h.strip()
    if h_clean in NARROW_COLS:
        GRID_WIDTHS.append(80)
    elif h_clean in MEDIUM_COLS:
        GRID_WIDTHS.append(140)
    elif h_clean in WIDE_COLS:
        GRID_WIDTHS.append(250)
    else:
        GRID_WIDTHS.append(100)


def get_base_path() -> Path:
    return Path(sys._MEIPASS) if getattr(sys, 'frozen', False) else Path(__file__).parent.parent

def _brand_button(parent, text: str, command, state=tk.NORMAL, bg_color=_PRIMARY_BLUE, secondary=False) -> tk.Button:
    if secondary:
        btn = tk.Button(parent, text=text, command=command, state=state,
                        font=("Segoe UI",10,"bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE,
                        activebackground=_LIGHT_BG, activeforeground=_PRIMARY_BLUE,
                        bd=1, relief=tk.SOLID, highlightthickness=0, highlightbackground=_PRIMARY_BLUE,
                        padx=14, pady=5, cursor="hand2")
    else:
        btn = tk.Button(parent, text=text, command=command, state=state,
                        font=("Segoe UI",10,"bold"), fg="#FFF", bg=bg_color,
                        activebackground=_HOVER_BLUE, activeforeground="#FFF",
                        bd=0, padx=14, pady=6, cursor="hand2", relief=tk.FLAT)
        if bg_color == _PRIMARY_BLUE:
            btn.bind("<Enter>", lambda e: btn.configure(bg=_HOVER_BLUE) if btn['state'] != 'disabled' else None)
            btn.bind("<Leave>", lambda e: btn.configure(bg=_PRIMARY_BLUE) if btn['state'] != 'disabled' else None)
    return btn

def align_sheet_columns(sheet, headers):
    for idx, header in enumerate(headers):
        sheet.align_columns(columns=idx, align="center")

def get_flexible(d, key):
    """Retrieves dict values matching key case-insensitively and stripped of whitespace."""
    if key in d:
        return d[key]
    key_clean = str(key).strip().lower()
    for k, v in d.items():
        if str(k).strip().lower() == key_clean:
            return v
    return ""


class ChecklistApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Nagarkot Checklist Template Generator")
        self.root.configure(bg=_LIGHT_BG)
        self.root.state("zoomed")
        self.root.minsize(1024, 600)
        
        # Configure standard TTK styles
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('TEntry', padding=6, font=("Segoe UI", 10), bordercolor=_BORDER_GRAY, lightcolor=_BORDER_GRAY, darkcolor=_BORDER_GRAY, fieldbackground=_PANEL_WHITE)
        self.style.configure('TCombobox', padding=6, font=("Segoe UI", 10), bordercolor=_BORDER_GRAY, lightcolor=_BORDER_GRAY, darkcolor=_BORDER_GRAY, fieldbackground=_PANEL_WHITE)
        self.style.configure('TNotebook', background=_LIGHT_BG)
        self.style.configure('TNotebook.Tab', padding=(16, 6), font=("Segoe UI", 10, "bold"), background=_PANEL_WHITE, foreground=_MUTED_GRAY)
        self.style.map('TNotebook.Tab', 
                       background=[('selected', _PRIMARY_BLUE)], 
                       foreground=[('selected', '#FFFFFF')],
                       padding=[('selected', (16, 6))])
        
        self.gsheets_url = tk.StringVar(value="https://docs.google.com/spreadsheets/d/1nZyxJGCwG7bK8-wg2Tti9l1C3sS5KcULefBvquW3u8E/edit?gid=1228640769#gid=1228640769")
        self.cred_path = tk.StringVar(value=str(get_base_path() / "credentials.json"))
        
        self.importer = None
        self.importer_df = None # Cached importer DataFrame
        self.rows_data = [] # List of dicts representing each added model row
        
        # 0-based index configs for tksheet
        self.editable_cols_0based = {i - 1 for i in EDITABLE_INDICES}
        self.readonly_cols_0based = [i for i in range(len(checklist_logic.LOGISYS_HEADERS)) if i not in self.editable_cols_0based]
        
        # Build main container
        self._build_header()
        
        self.body_container = tk.Frame(self.root, bg=_LIGHT_BG)
        self.body_container.pack(fill=tk.BOTH, expand=True, padx=16, pady=(12,4))
        
        self._build_footer()
        
        self.show_login_screen()

    def _build_header(self) -> None:
        header = tk.Frame(self.root, bg=_PANEL_WHITE, bd=0, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        header.pack(fill=tk.X, side=tk.TOP); header.pack_propagate(False); header.configure(height=64)
        
        # Logo Left
        left_frame = tk.Frame(header, bg=_PANEL_WHITE)
        left_frame.pack(side=tk.LEFT, fill=tk.Y)
        try:
            self._logo_img = tk.PhotoImage(file=str(get_base_path() / "logo.png"))
            factor = max(1, self._logo_img.height() // 20)
            self._logo_img = self._logo_img.subsample(factor)
            tk.Label(left_frame, image=self._logo_img, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=24, pady=22)
        except Exception:
            tk.Label(left_frame, text="Nagarkot", font=("Segoe UI",12,"bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=24, pady=22)
            
        # Absolute Center Title
        tf = tk.Frame(header, bg=_PANEL_WHITE); tf.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        tk.Label(tf, text="Checklist Template Generator", font=("Segoe UI",14,"bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE).pack()
        tk.Label(tf, text="Nagarkot Forwarders Pvt Ltd", font=("Segoe UI",9), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack()

        # Right Actions
        right_frame = tk.Frame(header, bg=_PANEL_WHITE)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=24)
        
        self.logout_btn = tk.Button(right_frame, text="Logout", command=self.show_login_screen, bg=_PANEL_WHITE, fg=_ACCENT_RED, bd=0, font=("Segoe UI", 10, "bold"), cursor="hand2")
        self.logout_btn.pack(side=tk.RIGHT, pady=18, padx=(10, 0))
        self.logout_btn.bind("<Enter>", lambda e: self.logout_btn.configure(bg="#FEE2E2"))
        self.logout_btn.bind("<Leave>", lambda e: self.logout_btn.configure(bg=_PANEL_WHITE))
        self.logout_btn.pack_forget()
        
        self.user_label = tk.Label(right_frame, text="", font=("Segoe UI", 10, "italic"), fg=_MUTED_GRAY, bg=_PANEL_WHITE)
        self.user_label.pack(side=tk.RIGHT, pady=18)
        self.user_label.pack_forget()

    def _build_footer(self) -> None:
        ft = tk.Frame(self.root, bg=_PANEL_WHITE, height=28, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        ft.pack(fill=tk.X, side=tk.BOTTOM); ft.pack_propagate(False)
        tk.Label(ft, text="Nagarkot Forwarders Pvt. Ltd. ©", font=("Segoe UI",8), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=12)

    def _clear_body(self):
        for widget in self.body_container.winfo_children():
            widget.destroy()

    # --- SCREEN 1: LOGIN ---
    def show_login_screen(self):
        self.logout_btn.pack_forget()
        if hasattr(self, 'user_label'):
            self.user_label.pack_forget()
        self._clear_body()
        auth.clear_session()
        
        login_frame = tk.Frame(self.body_container, bg=_PANEL_WHITE, padx=40, pady=40, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        login_frame.place(relx=0.5, rely=0.4, anchor=tk.CENTER)
        
        tk.Label(login_frame, text="System Login", font=("Segoe UI", 16, "bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE).pack(pady=(0, 20))
        
        # Credentials & URL Setup
        tk.Label(login_frame, text="GSheets Master URL:", bg=_PANEL_WHITE, font=("Segoe UI", 9)).pack(anchor='w')
        ttk.Entry(login_frame, textvariable=self.gsheets_url, width=50).pack(pady=(0, 10))
        
        tk.Label(login_frame, text="Credentials JSON:", bg=_PANEL_WHITE, font=("Segoe UI", 9)).pack(anchor='w')
        cred_entry = ttk.Entry(login_frame, textvariable=self.cred_path, width=50)
        cred_entry.pack(pady=(0, 10))
        
        tk.Label(login_frame, text="Username:", bg=_PANEL_WHITE, font=("Segoe UI", 9)).pack(anchor='w')
        username_var = tk.StringVar()
        ttk.Entry(login_frame, textvariable=username_var, width=50).pack(pady=(0, 10))
        
        tk.Label(login_frame, text="Password:", bg=_PANEL_WHITE, font=("Segoe UI", 9)).pack(anchor='w')
        password_var = tk.StringVar()
        ttk.Entry(login_frame, textvariable=password_var, width=50, show="*").pack(pady=(0, 20))
        
        def do_login():
            url = self.gsheets_url.get().strip()
            if not url:
                messagebox.showerror("Error", "GSheets URL is required.")
                return
            
            self._clear_body()
            loading_label = tk.Label(self.body_container, text="⌛ Logging in and fetching importer list...", 
                                     font=("Segoe UI", 12, "bold"), fg=_PRIMARY_BLUE, bg=_LIGHT_BG)
            loading_label.place(relx=0.5, rely=0.4, anchor=tk.CENTER)
            self.root.update()
            
            def auth_thread():
                try:
                    success, msg = auth.authenticate_user(url, self.cred_path.get(), username_var.get().strip(), password_var.get().strip())
                    if success:
                        from gsheets import GSheetsClient
                        client = GSheetsClient(self.cred_path.get())
                        sheet_id = client.extract_sheet_id(self.gsheets_url.get())
                        spreadsheet = client.client.open_by_key(sheet_id)
                        self.tab_names = [ws.title for ws in spreadsheet.worksheets() if ws.title not in ['Tool_Users', 'Audit_Log', 'Invoice_Header_Mappings', 'Pending_Model_Approval']]
                        
                        try:
                            self.invoice_mappings_df, _ = client.get_sheet_data(self.gsheets_url.get(), "Invoice_Header_Mappings")
                        except Exception as mapping_err:
                            print(f"Warning: Could not load Invoice_Header_Mappings. {mapping_err}")
                            self.invoice_mappings_df = None
                            
                        self.importer = None
                        self.importer_df = None
                        self.rows_data = []
                        self.root.after(0, self.show_builder_screen)
                    else:
                        self.root.after(0, lambda m=msg: handle_login_fail(m))
                except Exception as e:
                    self.root.after(0, lambda err=str(e): handle_login_fail(err))
                    
            def handle_login_fail(err_msg):
                messagebox.showerror("Login Failed", err_msg)
                self.show_login_screen()
                
            threading.Thread(target=auth_thread, daemon=True).start()
                
        _brand_button(login_frame, "Login", do_login).pack(fill=tk.X)



    # --- SCREEN 3: CHECKLIST BUILDER ---
    def show_builder_screen(self):
        self.logout_btn.pack(side=tk.RIGHT, pady=18, padx=(10, 0))
        if hasattr(self, 'user_label'):
            session = auth.get_current_session()
            if session and isinstance(session, dict):
                username = session.get('Username', 'User')
                self.user_label.config(text=f"👤 {username}")
                self.user_label.pack(side=tk.RIGHT, pady=18)
        self._clear_body()
        
        def _do_clear():
            if self.rows_data and messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all rows?"):
                self.rows_data = []
                self.session_queue = []
                self._reload_table_rows()

        # Control Panel Wrapper
        ctrl_wrapper = tk.Frame(self.body_container, bg=_PANEL_WHITE, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        ctrl_wrapper.pack(fill=tk.X, pady=(0, 16))
        
        # Left Side: Importer
        left_ctrl = tk.Frame(ctrl_wrapper, bg=_PANEL_WHITE, padx=16, pady=16)
        left_ctrl.pack(side=tk.LEFT, fill=tk.Y)
        
        tk.Label(left_ctrl, text="Importer:", font=("Segoe UI", 9, "bold"), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack(anchor='w', pady=(0,4))
        
        importer_combo = ttk.Combobox(left_ctrl, width=25)
        if hasattr(self, 'tab_names') and self.tab_names:
            importer_combo['values'] = self.tab_names
            if self.importer in self.tab_names:
                importer_combo.set(self.importer)
            else:
                importer_combo.set("Select Importer...")
        else:
            importer_combo.set("Select Importer...")
        importer_combo.pack(anchor='w')
        
        # Third Column: Search Master Sheet
        search_ctrl = tk.Frame(ctrl_wrapper, bg=_PANEL_WHITE, padx=16, pady=16)
        search_ctrl.pack(side=tk.RIGHT, fill=tk.Y)
        
        tk.Label(search_ctrl, text="Search Master Sheet:", font=("Segoe UI", 9, "bold"), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack(anchor='w', pady=(0,4))
        
        search_input_row = tk.Frame(search_ctrl, bg=_PANEL_WHITE)
        search_input_row.pack(fill=tk.X)
        
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_input_row, textvariable=self.search_var, width=25)
        search_entry.pack(side=tk.LEFT, padx=(0,10))
        
        # We define _do_master_search later or pass a lambda that calls it
        search_btn = _brand_button(search_input_row, "🔍 Search", lambda: self._do_master_search())
        search_btn.pack(side=tk.LEFT)
        search_entry.bind('<Return>', lambda e: self._do_master_search())

        # Middle Column: Actions
        right_ctrl = tk.Frame(ctrl_wrapper, bg=_PANEL_WHITE, padx=16, pady=16)
        right_ctrl.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(right_ctrl, text="Add Models (comma-separated):", font=("Segoe UI", 9, "bold"), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack(anchor='w', pady=(0,4))

        input_row = tk.Frame(right_ctrl, bg=_PANEL_WHITE)
        input_row.pack(fill=tk.X)

        models_var = tk.StringVar()
        models_entry = ttk.Entry(input_row, textvariable=models_var)
        models_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # We will pack the action buttons into input_row later below.
        
        status_row = tk.Frame(right_ctrl, bg=_PANEL_WHITE)
        status_row.pack(fill=tk.X, pady=(4,0))
        
        def filter_builder_tabs(event):
            if event.keysym in ("Up", "Down", "Left", "Right", "Return", "Escape"):
                return
            typed = importer_combo.get().strip()
            if not typed:
                importer_combo['values'] = self.tab_names
            else:
                filtered = [t for t in self.tab_names if typed.lower() in t.lower()]
                importer_combo['values'] = filtered
        importer_combo.bind('<KeyRelease>', filter_builder_tabs)
        
        def on_importer_changed(event):
            new_imp = importer_combo.get().strip()
            if not new_imp:
                return
            if hasattr(self, 'tab_names') and self.tab_names:
                if new_imp not in self.tab_names:
                    matches = [t for t in self.tab_names if new_imp.lower() in t.lower()]
                    if len(matches) == 1:
                        new_imp = matches[0]
                    else:
                        messagebox.showerror("Error", f"Invalid importer name: '{new_imp}'")
                        importer_combo.set(self.importer)
                        return
            if new_imp and new_imp != self.importer:
                if self.rows_data:
                    if not messagebox.askyesno("Confirm", "Changing importer will clear your current checklist rows. Proceed?"):
                        importer_combo.set(self.importer)
                        return
                
                self.importer = new_imp
                self._clear_body()
                loading_imp = tk.Label(self.body_container, text=f"⌛ Loading master database for '{new_imp}'...", 
                                       font=("Segoe UI", 12, "bold"), fg=_PRIMARY_BLUE, bg=_LIGHT_BG)
                loading_imp.place(relx=0.5, rely=0.4, anchor=tk.CENTER)
                self.root.update()
                
                def load_df_thread():
                    try:
                        from gsheets import GSheetsClient
                        client = GSheetsClient(self.cred_path.get())
                        df, _ = client.get_sheet_data(self.gsheets_url.get(), new_imp)
                        if not df.empty:
                            for col in df.columns:
                                if str(col).strip().lower() == 'model':
                                    df.rename(columns={col: 'Model'}, inplace=True)
                                    break
                            if 'Model' in df.columns:
                                df['Model'] = df['Model'].astype(str).str.strip()
                        self.importer_df = df
                        self.rows_data = [] # Reset rows
                        self.session_queue = []
                        self.root.after(0, self.show_builder_screen)
                    except Exception as e:
                        err_str = str(e)
                        def handle_error(err):
                            messagebox.showerror("Error", f"Failed to load data:\n{err}")
                            self.importer = None
                            self.importer_df = None
                            self.show_builder_screen()
                        self.root.after(0, lambda err=err_str: handle_error(err))
                        
                threading.Thread(target=load_df_thread, daemon=True).start()
                
        importer_combo.bind("<<ComboboxSelected>>", on_importer_changed)
        importer_combo.bind("<Return>", on_importer_changed)
        
        def do_add_models():
            if not self.importer or self.importer_df is None:
                messagebox.showerror("Error", "Please select an importer first.")
                return
            input_str = models_var.get().strip()
            if not input_str:
                messagebox.showerror("Error", "Please enter at least one model.")
                return
                
            models_to_lookup = [m.strip() for m in input_str.split(',') if m.strip()]
            
            if not hasattr(self, 'session_queue'):
                self.session_queue = []
            self.session_queue.extend(models_to_lookup)
            
            self.root.config(cursor="watch")
            self._process_model_queue(models_to_lookup)
            models_var.set("") # Clear entry
            
        self.status_label = tk.Label(status_row, text="", font=("Segoe UI", 9, "italic"), bg=_PANEL_WHITE, fg=_MUTED_GRAY)
        
        def _handle_upload_err(err):
            self.root.config(cursor="")
            self.status_label.config(text="")
            messagebox.showerror("Error", f"Failed to read Excel file: {err}")
            
        def _do_upload_invoice():
            if not self.importer or self.importer_df is None:
                messagebox.showerror("Error", "Please select an importer first.")
                return
            if getattr(self, 'invoice_mappings_df', None) is None:
                messagebox.showerror("Error", "Invoice mapping configuration could not be loaded. Try relogging.")
                return
                
            file_path = filedialog.askopenfilename(
                title="Select Invoice",
                filetypes=[("Excel & PDF Files", "*.xlsx *.xls *.pdf")]
            )
            if not file_path:
                return
                
            self.status_label.config(text="⏳ Reading file...")
            self.root.config(cursor="watch")
            self.root.update_idletasks()
            
            def parse_thread():
                try:
                    import pandas as pd
                    if self.importer.lower() == 'advics':
                        invoice_df = pd.read_excel(file_path, header=None)
                        self.root.after(0, lambda: _parse_advics_invoice(invoice_df))
                    elif self.importer.lower() == 'arjo':
                        invoice_df = pd.read_excel(file_path, header=None)
                        self.root.after(0, lambda: _parse_arjo_invoice(invoice_df))
                    elif self.importer.lower() == 'ansell':
                        if not file_path.lower().endswith('.pdf'):
                            raise ValueError("Ansell importer requires a PDF invoice.")
                        self.root.after(0, lambda: _parse_ansell_invoice(file_path))
                    elif self.importer.lower() == 'aviat':
                        if file_path.lower().endswith('.pdf'):
                            self.root.after(0, lambda: _parse_aviat_invoice(file_path))
                        else:
                            invoice_df = pd.read_excel(file_path)
                            self.root.after(0, lambda: _parse_aviat_excel(invoice_df))
                    elif self.importer.lower() in ['bharti airtel', 'bharti hexacom']:
                        if file_path.lower().endswith('.pdf'):
                            self.root.after(0, lambda: _parse_airtel_invoice(file_path))
                        else:
                            invoice_df = pd.read_excel(file_path)
                            self.root.after(0, lambda: _on_file_parsed(invoice_df))
                    else:
                        invoice_df = pd.read_excel(file_path)
                        self.root.after(0, lambda: _on_file_parsed(invoice_df))
                except Exception as e:
                    self.root.after(0, lambda err=e: _handle_upload_err(err))
                    
            threading.Thread(target=parse_thread, daemon=True).start()
            
        def _parse_airtel_invoice(file_path):
            self.root.config(cursor="")
            self.status_label.config(text="")
            
            # Detect format (Ciena vs ECI vs Ceragon)
            pdf_type = "unknown"
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    first_page_text = pdf.pages[0].extract_text() or ""
                    first_lower = first_page_text.lower()
                    if "ciena" in first_lower:
                        pdf_type = "ciena"
                    elif "eci telecom" in first_lower:
                        pdf_type = "eci"
                    elif "ceragon" in first_lower:
                        pdf_type = "ceragon"
            except Exception:
                pass
            
            if self.importer.lower() == 'bharti hexacom' and pdf_type != 'ceragon':
                messagebox.showerror(
                    "Unrecognized Format",
                    "Bharti Hexacom only supports the Ceragon invoice format.\n\n"
                    "Please upload a valid Ceragon PDF invoice."
                )
                return
                
            if pdf_type == "unknown":
                messagebox.showerror(
                    "Unrecognized Format",
                    "This PDF does not match any known Airtel invoice format "
                    "(Ceragon, Ciena, or ECI Telecom).\n\n"
                    "Please check the file and try again, or use manual model entry."
                )
                return
                
            try:
                if pdf_type == "ciena":
                    import airtel_ciena_extractor
                    inv_no, doc_date, currency, items = airtel_ciena_extractor.parse_ciena_invoice(file_path)
                elif pdf_type == "eci":
                    import airtel_eci_extractor
                    inv_no, doc_date, currency, items = airtel_eci_extractor.parse_eci_invoice(file_path)
                else:
                    import airtel_ceragon_extractor
                    inv_no, doc_date, currency, items = airtel_ceragon_extractor.parse_ceragon_invoice(file_path)
            except Exception as e:
                messagebox.showerror("Parse Error", f"Failed to parse Airtel {pdf_type.capitalize()} PDF:\n{str(e)}")
                return
                
            if not items:
                messagebox.showinfo("Info", "No line items found in the PDF.")
                return
                
            target_to_col_idx = {}
            for t_key in ["Inv No", "Date", "Quantity", "Rate", "Country of origin", "Currency", "CETH", "Unit", "CTH", "Product Desc"]:
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
                        
            queue_items = []
            for item in items:
                model_val = item.get("part_no", "")
                if not model_val: continue
                
                if pdf_type in ("ciena", "eci"):
                    prefilled = {
                        target_to_col_idx.get("Inv No"): inv_no,
                        target_to_col_idx.get("Date"): doc_date,
                        target_to_col_idx.get("Quantity"): item.get("quantity", ""),
                        target_to_col_idx.get("Rate"): item.get("rate", ""),
                        target_to_col_idx.get("Country of origin"): item.get("coo", ""),
                        target_to_col_idx.get("Currency"): currency,
                        target_to_col_idx.get("Unit"): item.get("unit", ""),
                        target_to_col_idx.get("CTH"): item.get("cth", ""),
                        target_to_col_idx.get("Product Desc"): item.get("description", "")
                    }
                else:
                    prefilled = {
                        target_to_col_idx.get("Inv No"): inv_no,
                        target_to_col_idx.get("Date"): doc_date,
                        target_to_col_idx.get("Quantity"): item.get("quantity", ""),
                        target_to_col_idx.get("Rate"): item.get("unit_price", ""),
                        target_to_col_idx.get("Country of origin"): item.get("coo", ""),
                        target_to_col_idx.get("Currency"): currency,
                        target_to_col_idx.get("CETH"): "NOEXCISE",
                        target_to_col_idx.get("Unit"): "NOS"
                    }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})
                
            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            
        def _parse_aviat_excel(df):
            self.root.config(cursor="")
            self.status_label.config(text="")
            
            # Check for required columns based on user instruction
            req_cols = ["Commercial Invoice No", "Document Date (Long)", "Part No", "Quantity", "Unit Price"]
            missing_cols = [c for c in req_cols if c not in df.columns]
            if missing_cols:
                messagebox.showerror("Error", f"Missing required columns in Excel: {', '.join(missing_cols)}")
                return
                
            if "CoO" not in df.columns:
                messagebox.showerror("Missing Column", "Country of origin is blank.\n\nPlease add a 'CoO' column to the excel file and enter the CoO.")
                return
                
            if "Currency" not in df.columns:
                messagebox.showerror("Missing Column", "Currency column is missing.\n\nPlease add a 'Currency' column to the excel file and add details in it.")
                return
                
            target_to_col_idx = {}
            mapping = {
                "Inv No": "col_1", "Date": "col_2",
                "Quantity": "col_5", "Unit": "col_6", "Currency": "col_7",
                "Rate": "col_8", "Country of origin": "col_32"
            }
            # Add missing mappings dynamically
            import pandas as pd
            for t_key in mapping.keys():
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
                        
            queue_items = []
            for _, row in df.iterrows():
                model_val = str(row.get("Part No", "")).strip()
                if not model_val or pd.isna(row.get("Part No")) or model_val.lower() == 'nan':
                    continue
                    
                inv_no = str(row.get("Commercial Invoice No", "")).strip()
                if pd.isna(row.get("Commercial Invoice No")) or inv_no.lower() == 'nan':
                    inv_no = ""
                    
                date_val = row.get("Document Date (Long)", "")
                if hasattr(date_val, 'strftime'):
                    date_val = date_val.strftime('%d-%m-%Y')
                else:
                    date_val = str(date_val).strip()
                    if date_val.lower() == 'nan': date_val = ""
                    
                qty = str(row.get("Quantity", "")).strip()
                if qty.lower() == 'nan': qty = ""
                
                rate = str(row.get("Unit Price", "")).strip()
                if rate.lower() == 'nan': rate = ""
                
                coo = str(row.get("CoO", "")).strip()
                if pd.isna(row.get("CoO")) or coo.lower() == 'nan' or not coo:
                    messagebox.showerror("Missing CoO", f"Country of origin is blank for part {model_val}.\n\nPlease add a 'CoO' column to the excel file and enter the CoO.")
                    return
                    
                currency = str(row.get("Currency", "")).strip()
                if pd.isna(row.get("Currency")) or currency.lower() == 'nan' or not currency:
                    messagebox.showerror("Missing Currency", f"Currency is blank for part {model_val}.\n\nPlease add details in the Currency column.")
                    return
                
                prefilled = {
                    target_to_col_idx.get("Inv No"): inv_no,
                    target_to_col_idx.get("Date"): date_val,
                    target_to_col_idx.get("Quantity"): qty,
                    target_to_col_idx.get("Rate"): rate,
                    target_to_col_idx.get("Country of origin"): coo,
                    target_to_col_idx.get("Unit"): "NOS",
                    target_to_col_idx.get("Currency"): currency
                }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})
                
            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                messagebox.showinfo("Info", "No valid rows found in the Aviat Excel.")
            
        def _parse_aviat_invoice(pdf_path):
            try:
                import sys, os
                sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
                from aviat_inv_extractor import parse_commercial_invoice, format_date_helper, trim_hs_code
            except ImportError as e:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"Failed to load Aviat extractor: {e}")
                return
                
            try:
                inv_no, date_raw, ci_items = parse_commercial_invoice(pdf_path)
            except Exception as e:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"Failed to parse Aviat PDF: {e}")
                return
                
            if not ci_items:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid rows found in the Aviat invoice.")
                return
                
            # Currency detection: look below the word 'CURRENCY'
            currency_val = ""
            try:
                import pdfplumber
                with pdfplumber.open(pdf_path) as pdf:
                    first_page_text = pdf.pages[0].extract_text() or ""
                    lines = first_page_text.split('\n')
                    for i, line in enumerate(lines):
                        if "CURRENCY" in line.upper():
                            if i + 1 < len(lines):
                                next_line = lines[i+1].strip()
                                parts = next_line.split()
                                if parts:
                                    currency_val = parts[-1] # The currency code is usually the last word
                            break
            except Exception:
                pass
                
            date_short, date_long = format_date_helper(date_raw)
            
            target_to_col_idx = {}
            mapping = {
                "Inv No": "col_1", "Date": "col_2", "Product Desc": "col_4", 
                "Quantity": "col_5", "Unit": "col_6", "Currency": "col_7", 
                "Rate": "col_8", "Country of origin": "col_32", "CT (HS Code)": "col_42"
            }
            # Add missing mappings dynamically
            for t_key in mapping.keys():
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
                        
            queue_items = []
            for item in ci_items:
                model_val = item.get("part_no", "")
                if not model_val or model_val == "N/A":
                    continue
                    
                desc_upper = (item.get("description") or "").upper()
                ct_trimmed = trim_hs_code(item.get("ct"))
                
                prefilled = {
                    target_to_col_idx.get("Inv No"): inv_no,
                    target_to_col_idx.get("Date"): date_short,
                    target_to_col_idx.get("Product Desc"): desc_upper,
                    target_to_col_idx.get("Quantity"): str(item.get("quantity", "")),
                    target_to_col_idx.get("Unit"): "NOS",
                    target_to_col_idx.get("Currency"): currency_val,
                    target_to_col_idx.get("Rate"): str(item.get("unit_price", "")),
                    target_to_col_idx.get("Country of origin"): item.get("coo", ""),
                    target_to_col_idx.get("CT (HS Code)"): ct_trimmed,
                }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})
                
            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid models found in the Aviat invoice.")
            
        def _parse_ansell_invoice(pdf_path):
            try:
                import pdfplumber
            except ImportError:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", "pdfplumber is not installed.")
                return
                
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    all_tables = []
                    for page in pdf.pages:
                        all_tables.extend(page.extract_tables())
            except Exception as e:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"Failed to read PDF: {e}")
                return

            inv_table = None
            for tbl in all_tables:
                if len(tbl) >= 3 and tbl[0] and tbl[0][0] and "Commercial Invoice" in tbl[0][0]:
                    inv_table = tbl
                    break

            if inv_table is None:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"Invoice header table not found in {pdf_path}")
                return

            invoice_no = inv_table[2][0].strip() if inv_table[2][0] else ""
            invoice_date = ""
            if len(inv_table[2]) > 1 and inv_table[2][1]:
                invoice_date = inv_table[2][1].strip()

            item_tables = []
            for tbl in all_tables:
                if not tbl or not tbl[0]: continue
                if any(cell and "Product Description" in cell for cell in tbl[0]):
                    item_tables.append(tbl)

            if not item_tables:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"No item tables found in {pdf_path}")
                return

            def find_col(keyword, rows):
                kw = keyword.lower()
                for row in rows:
                    if not row: continue
                    for idx, cell in enumerate(row):
                        if cell and kw in cell.lower(): return idx
                return None

            records = []
            for tbl in item_tables:
                hdr = tbl[0]
                sub = tbl[1] if len(tbl) > 1 else [None] * len(hdr)

                idx_prod_code = find_col("product", [hdr])
                if idx_prod_code is not None and hdr[idx_prod_code] and "Description" in hdr[idx_prod_code]:
                    for idx, cell in enumerate(hdr):
                        if cell and "Product" in cell and "Code" in cell:
                            idx_prod_code = idx
                            break

                idx_desc = find_col("Product Description", [hdr])
                idx_country = find_col("Country", [hdr])
                
                idx_hs = find_col("HTS/HS", [hdr])
                idx_shipcase = find_col("Case", [hdr])
                idx_value = find_col("Value", [hdr])

                idx_qty = None
                idx_uom = None
                idx_price = None
                for idx, cell in enumerate(sub):
                    if cell == "Qty" and idx_qty is None: idx_qty = idx
                    elif cell == "UOM" and idx_qty is not None and idx_uom is None: idx_uom = idx
                    elif cell == "Price" and idx_price is None: idx_price = idx

                idx_currency = find_col("Cur", [hdr])

                def get(row, idx):
                    if idx is None or idx >= len(row): return ""
                    return (row[idx] or "").strip()

                for row in tbl[2:]:
                    if not row: continue
                    first_cell = (row[0] or "").strip() if len(row) > 0 else ""
                    if first_cell.startswith("Total"): break
                    if first_cell == "": continue

                    record = {
                        "Invoice_No": invoice_no,
                        "Invoice_Date": invoice_date,
                        "Product_Code": get(row, idx_prod_code),
                        "Product_Description": get(row, idx_desc),
                        "Country_of_Origin": get(row, idx_country),
                        "Qty": get(row, idx_qty),
                        "UOM": get(row, idx_uom),
                        "Price": get(row, idx_price),
                        "Currency": get(row, idx_currency),
                        "HS_Code": get(row, idx_hs),
                        "Ship_Case": get(row, idx_shipcase),
                        "Value": get(row, idx_value),
                    }
                    records.append(record)
                    
            target_to_col_idx = {}
            mapping = {
                "Inv No": "col_1", "Date": "col_2", "Product Desc": "col_4", 
                "Quantity": "col_5", "Unit": "col_6", "Currency": "col_7", 
                "Rate": "col_8", "Country of origin": "col_32"
            }
            for t_key in mapping.keys():
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
                        
            queue_items = []
            for rec in records:
                model_val = rec.get("Product_Code", "")
                if not model_val:
                    continue
                    
                # Rate calculation: Value / Qty
                ansell_rate = ""
                qty_val = rec.get("Qty", "")
                val_val = rec.get("Value", "")
                if qty_val and val_val:
                    try:
                        q_num = float(qty_val.replace(',', ''))
                        v_num = float(val_val.replace(',', ''))
                        if q_num > 0:
                            ansell_rate = str(round(v_num / q_num, 4))
                    except ValueError:
                        pass
                        
                prefilled = {
                    target_to_col_idx.get("Inv No"): rec.get("Invoice_No", ""),
                    target_to_col_idx.get("Date"): rec.get("Invoice_Date", ""),
                    target_to_col_idx.get("Quantity"): rec.get("Qty", ""),
                    target_to_col_idx.get("Unit"): rec.get("UOM", ""),
                    target_to_col_idx.get("Currency"): rec.get("Currency", ""),
                    target_to_col_idx.get("Rate"): ansell_rate,
                    target_to_col_idx.get("Country of origin"): rec.get("Country_of_Origin", ""),
                    
                    # Internal keys for Ansell description construction
                    "_ansell_raw_desc": rec.get("Product_Description", ""),
                    "_ansell_hs": rec.get("HS_Code", ""),
                    "_ansell_case": rec.get("Ship_Case", ""),
                    "_ansell_unit_price": rec.get("Price", "")
                }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})

            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid rows found in the Ansell invoice.")
                
        def _parse_advics_invoice(invoice_df):
            import pandas as pd
            inv_no, inv_date, coo_val = "", "", ""
            
            for r in range(min(30, len(invoice_df))):
                for c in range(len(invoice_df.columns)):
                    val = str(invoice_df.iat[r, c]).strip()
                    if val == "INVOICE NO. :":
                        for c_next in range(c + 1, len(invoice_df.columns)):
                            if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                inv_no = str(invoice_df.iat[r, c_next]).strip()
                                break
                    elif val == "DATE :":
                        for c_next in range(c + 1, len(invoice_df.columns)):
                            if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                d_val = invoice_df.iat[r, c_next]
                                if hasattr(d_val, 'strftime'):
                                    inv_date = d_val.strftime('%d-%m-%Y')
                                else:
                                    inv_date = str(d_val).strip()
                                break
                    elif val == "Country of origin :":
                        for c_next in range(c + 1, len(invoice_df.columns)):
                            if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                coo_val = str(invoice_df.iat[r, c_next]).strip()
                                break

            table_start_row = -1
            model_col, desc_col, qty_col, price_col = -1, -1, -1, -1
            currency = "THB"
            
            for r in range(len(invoice_df)):
                row_vals = [str(x).strip() for x in invoice_df.iloc[r] if pd.notna(x)]
                if "PART NO." in row_vals and "DESCRIPTION" in row_vals:
                    table_start_row = r
                    for c in range(len(invoice_df.columns)):
                        val = str(invoice_df.iat[r, c]).strip()
                        if val == "PART NO.": model_col = c
                        elif val == "DESCRIPTION": desc_col = c
                        elif val == "Qty": qty_col = c
                        elif "Unit Price" in val:
                            price_col = c
                            if "(THB)" in val: currency = "THB"
                    break
                    
            if table_start_row == -1 or model_col == -1:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror(
                    "Missing Required Headers",
                    "Could not find required item table headers in the Advics invoice.\n\n"
                    "The tool was searching for the following column names:\n"
                    "  • Model / Part No: 'PART NO.'\n"
                    "  • Description: 'DESCRIPTION'\n"
                    "  • Quantity: 'Qty'\n"
                    "  • Unit Price: 'Unit Price' (or 'Unit Price (THB)')\n\n"
                    "Please check your Excel file and ensure these header names are present."
                )
                return

            target_to_col_idx = {}
            mapping = {
                "Inv No": "col_1", "Date": "col_2", "Product Desc": "col_4", 
                "Quantity": "col_5", "Unit": "col_6", "Currency": "col_7", 
                "Rate": "col_8", "Country of origin": "col_32"
            }
            for t_key, expected_col in mapping.items():
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
            
            queue_items = []
            for r in range(table_start_row + 1, len(invoice_df)):
                if pd.isna(invoice_df.iat[r, model_col]):
                    continue
                    
                model_val = str(invoice_df.iat[r, model_col]).strip()
                if not model_val or model_val.lower() == 'nan':
                    continue
                    
                desc_val = str(invoice_df.iat[r, desc_col]).strip() if desc_col != -1 and pd.notna(invoice_df.iat[r, desc_col]) else ""
                qty_val = str(invoice_df.iat[r, qty_col]).strip() if qty_col != -1 and pd.notna(invoice_df.iat[r, qty_col]) else ""
                price_val = str(invoice_df.iat[r, price_col]).strip() if price_col != -1 and pd.notna(invoice_df.iat[r, price_col]) else ""
                
                unit_val = ""
                if qty_col != -1:
                    for c_next in range(qty_col + 1, len(invoice_df.columns)):
                        c_val = str(invoice_df.iat[r, c_next]).strip()
                        if c_val and c_val.lower() != 'nan':
                            unit_val = c_val
                            break
                            
                if unit_val.lower() in ("pcs.", "pcs"):
                    unit_val = "NOS"
                
                prefilled = {
                    target_to_col_idx.get("Inv No"): inv_no,
                    target_to_col_idx.get("Date"): inv_date,
                    target_to_col_idx.get("Product Desc"): desc_val,
                    target_to_col_idx.get("Quantity"): qty_val,
                    target_to_col_idx.get("Unit"): unit_val,
                    target_to_col_idx.get("Currency"): currency,
                    target_to_col_idx.get("Rate"): price_val,
                    target_to_col_idx.get("Country of origin"): coo_val
                }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})

            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid rows found in the Advics invoice.")
                
        def _parse_arjo_invoice(invoice_df):
            import pandas as pd
            import re
            inv_no, inv_date, header_coo = "", "", ""
            
            for r in range(min(50, len(invoice_df))):
                for c in range(len(invoice_df.columns)):
                    val = str(invoice_df.iat[r, c]).strip()
                    if "Invoice No." in val or "Invoice No" in val:
                        match = re.search(r'Invoice No\.?\s*(\S+)', val, re.IGNORECASE)
                        if match and match.group(1).strip():
                            inv_no = match.group(1).strip()
                        else:
                            for c_next in range(c + 1, len(invoice_df.columns)):
                                if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                    inv_no = str(invoice_df.iat[r, c_next]).strip()
                                    break
                    elif "Invoice date :" in val or "Invoice date" in val:
                        match = re.search(r'Invoice date\s*:\s*(.*)', val, re.IGNORECASE)
                        d_val = ""
                        if match and match.group(1).strip():
                            d_val = match.group(1).strip()
                        else:
                            for c_next in range(c + 1, len(invoice_df.columns)):
                                if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                    d_val = invoice_df.iat[r, c_next]
                                    break
                        if d_val:
                            if hasattr(d_val, 'strftime'):
                                inv_date = d_val.strftime('%d-%m-%Y')
                            else:
                                try:
                                    parsed_date = pd.to_datetime(str(d_val).strip())
                                    inv_date = parsed_date.strftime('%d-%m-%Y')
                                except:
                                    inv_date = str(d_val).strip()
                    elif "Country Of Origin" in val or "Country of Origin" in val:
                        match = re.search(r'Country\s+Of\s+Origin\s*:\s*(.*)', val, re.IGNORECASE)
                        if match and match.group(1).strip():
                            header_coo = match.group(1).strip()
                        else:
                            for c_next in range(c + 1, len(invoice_df.columns)):
                                if pd.notna(invoice_df.iat[r, c_next]) and str(invoice_df.iat[r, c_next]).strip():
                                    header_coo = str(invoice_df.iat[r, c_next]).strip()
                                    break

            table_start_row = -1
            model_col, desc_col, qty_col, price_col, coo_col = -1, -1, -1, -1, -1
            currency = ""
            
            for r in range(len(invoice_df)):
                row_vals = [str(x).strip() for x in invoice_df.iloc[r] if pd.notna(x)]
                if "Part No" in row_vals and any(q in row_vals for q in ["Qty", "Quantity"]):
                    table_start_row = r
                    for c in range(len(invoice_df.columns)):
                        val = str(invoice_df.iat[r, c]).strip()
                        if val == "Part No": model_col = c
                        elif "Product Description" in val or "Product Desc" in val: desc_col = c
                        elif val in ("Qty", "Quantity"): qty_col = c
                        elif "Country Of Origin" in val or "Country of Origin" in val: coo_col = c
                        elif "Price" in val or "Rate" in val:
                            price_col = c
                            match = re.search(r'\((.*?)\)', val)
                            if match:
                                currency = match.group(1).strip()
                    break
                    
            if table_start_row == -1 or model_col == -1:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror(
                    "Missing Required Headers",
                    "Could not find required item table headers in the Arjo invoice.\n\n"
                    "The tool was searching for the following column names:\n"
                    "  • Model / Part No: 'Part No'\n"
                    "  • Quantity: 'Qty' or 'Quantity'\n"
                    "  • Description: 'Product Description' or 'Product Desc'\n"
                    "  • Price: 'Price' or 'Per Unit Price'\n\n"
                    "Please check your Excel file and ensure these header names are present."
                )
                return

            target_to_col_idx = {}
            mapping = {
                "Inv No": "col_1", "Date": "col_2", "Product Desc": "col_4",
                "Quantity": "col_5", "Unit": "col_6", "Currency": "col_7", 
                "Rate": "col_8", "Country of origin": "col_32"
            }
            for t_key, expected_col in mapping.items():
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_key.lower():
                        target_to_col_idx[t_key] = f"col_{i+1}"
                        break
            
            queue_items = []
            for r in range(table_start_row + 1, len(invoice_df)):
                if pd.isna(invoice_df.iat[r, model_col]):
                    continue
                    
                model_val = str(invoice_df.iat[r, model_col]).strip()
                if not model_val or model_val.lower() in ('nan', 'total', 'part no'):
                    continue
                if 'total' in model_val.lower():
                    break
                    
                qty_val = str(invoice_df.iat[r, qty_col]).strip() if qty_col != -1 and pd.notna(invoice_df.iat[r, qty_col]) else ""
                desc_val = str(invoice_df.iat[r, desc_col]).strip() if desc_col != -1 and pd.notna(invoice_df.iat[r, desc_col]) else ""
                price_val = str(invoice_df.iat[r, price_col]).strip() if price_col != -1 and pd.notna(invoice_df.iat[r, price_col]) else ""
                coo_val = str(invoice_df.iat[r, coo_col]).strip() if coo_col != -1 and pd.notna(invoice_df.iat[r, coo_col]) else ""
                if not coo_val:
                    coo_val = header_coo
                    
                if qty_val.lower() == 'nan': qty_val = ""
                if desc_val.lower() == 'nan': desc_val = ""
                if price_val.lower() == 'nan': price_val = ""
                if coo_val.lower() == 'nan': coo_val = ""
                
                unit_val = "NOS"
                
                prefilled = {
                    target_to_col_idx.get("Inv No"): inv_no,
                    target_to_col_idx.get("Date"): inv_date,
                    target_to_col_idx.get("Product Desc"): desc_val,
                    target_to_col_idx.get("Quantity"): qty_val,
                    target_to_col_idx.get("Unit"): unit_val,
                    target_to_col_idx.get("Currency"): currency,
                    target_to_col_idx.get("Rate"): price_val,
                    target_to_col_idx.get("Country of origin"): coo_val
                }
                prefilled = {k: v for k, v in prefilled.items() if k is not None}
                queue_items.append({'model': model_val, 'prefilled': prefilled})
                
            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid rows found in the Arjo invoice.")
            
        def _on_file_parsed(invoice_df):
            import pandas as pd
            importer_map = self.invoice_mappings_df[self.invoice_mappings_df['Importer'] == self.importer]
            
            variants = {}
            for _, row in importer_map.iterrows():
                fmt = str(row.get('Format_Name', '')).strip()
                if not fmt: continue
                sh = str(row.get('Source_Header', '')).strip()
                tf = str(row.get('Target_Field', '')).strip()
                if pd.notna(sh):
                    if fmt not in variants:
                        variants[fmt] = {}
                    variants[fmt][sh] = tf
                    
            excel_headers = [str(c).strip() for c in invoice_df.columns]
            excel_header_set = set(excel_headers)
            
            matched_variants = []
            for fmt, mapping in variants.items():
                if set(mapping.keys()) == excel_header_set:
                    matched_variants.append(fmt)
            
            # Superset match: if no exact match, check if file contains all
            # of a variant's headers (plus extras). This handles real invoices
            # that have extra columns we don't need.
            if not matched_variants:
                superset_matches = []
                for fmt, mapping in variants.items():
                    if set(mapping.keys()).issubset(excel_header_set):
                        superset_matches.append(fmt)
                if len(superset_matches) == 1:
                    matched_variants = superset_matches
                elif len(superset_matches) > 1:
                    # Multiple variants are subsets — pick the one with the
                    # most headers (most specific match) to break the tie
                    superset_matches.sort(key=lambda f: len(variants[f]), reverse=True)
                    if len(variants[superset_matches[0]]) > len(variants[superset_matches[1]]):
                        matched_variants = [superset_matches[0]]
                    # else: genuine ambiguity, fall through to guidance
                    
            expected_headers = {}
            self.active_variant_targets = set()
            
            if len(matched_variants) == 1:
                fmt = matched_variants[0]
                self.active_variant_targets = set(variants[fmt].values())
                for sh, tf in variants[fmt].items():
                    if pd.notna(tf) and tf:
                        expected_headers[sh] = tf
                        

            elif len(matched_variants) > 1:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", f"Ambiguous format match! Matched variants: {', '.join(matched_variants)}.")
                return
                
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                
                # Compute overlap scores to find the closest variant
                best_score = -1
                best_variants = []
                
                for fmt_name, fmt_mapping in variants.items():
                    variant_headers = set(fmt_mapping.keys())
                    overlap = len(variant_headers.intersection(excel_header_set))
                    if overlap > best_score:
                        best_score = overlap
                        best_variants = [(fmt_name, fmt_mapping)]
                    elif overlap == best_score:
                        best_variants.append((fmt_name, fmt_mapping))
                
                # Build the guidance dialog
                top = tk.Toplevel(self.root)
                top.title("Column Mismatch")
                top.grab_set()
                top.configure(bg="white")
                top.resizable(True, True)
                
                # Header bar
                header_frame = tk.Frame(top, bg="#FEF2F2", pady=8, padx=14)
                header_frame.pack(fill=tk.X)
                tk.Label(header_frame, text="⚠  Some column headers need renaming",
                         font=("Segoe UI", 11, "bold"), bg="#FEF2F2", fg="#991B1B").pack(anchor="w")
                tk.Label(header_frame, text="Rename these columns in your Excel file, save, then re-upload.",
                         font=("Segoe UI", 9), bg="#FEF2F2", fg="#7F1D1D").pack(anchor="w")
                
                # Scrollable content
                outer = tk.Frame(top, bg="white")
                outer.pack(fill=tk.BOTH, expand=True)
                canvas = tk.Canvas(outer, bg="white", highlightthickness=0)
                scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
                content = tk.Frame(canvas, bg="white", padx=16, pady=10)
                content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                canvas.create_window((0, 0), window=content, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)
                canvas.pack(side="left", fill="both", expand=True)
                scrollbar.pack(side="right", fill="y")
                
                def _build_single_variant_guidance(parent, fmt_mapping):
                    """Build guidance showing what's wrong and what to do, returns full header list."""
                    variant_headers = set(fmt_mapping.keys())
                    correct = sorted(variant_headers.intersection(excel_header_set))
                    missing = sorted(variant_headers - excel_header_set)
                    extra = sorted(excel_header_set - variant_headers)
                    
                    if missing:
                        # Section 1: What needs fixing
                        fix_frame = tk.Frame(parent, bg="#FEF2F2", padx=10, pady=8)
                        fix_frame.pack(fill=tk.X, pady=(0, 8))
                        
                        tk.Label(fix_frame, text="❌  Columns that need renaming:",
                                 font=("Segoe UI", 10, "bold"), bg="#FEF2F2", fg="#991B1B",
                                 anchor="w").pack(fill=tk.X, pady=(0, 6))
                        
                        # Table header
                        hdr_row = tk.Frame(fix_frame, bg="#FEF2F2")
                        hdr_row.pack(fill=tk.X, pady=(0, 3))
                        tk.Label(hdr_row, text="Your column", font=("Segoe UI", 8, "bold"),
                                 bg="#FEF2F2", fg="#6B7280", width=24, anchor="w").pack(side=tk.LEFT)
                        tk.Label(hdr_row, text="", bg="#FEF2F2", width=4).pack(side=tk.LEFT)
                        tk.Label(hdr_row, text="Rename to", font=("Segoe UI", 8, "bold"),
                                 bg="#FEF2F2", fg="#6B7280", width=24, anchor="w").pack(side=tk.LEFT)
                        
                        # Pair missing headers with extra (unrecognized) columns
                        for i, required_name in enumerate(missing):
                            pair_row = tk.Frame(fix_frame, bg="#FEF2F2")
                            pair_row.pack(fill=tk.X, pady=2)
                            
                            # Left side: user's likely wrong column (if one exists)
                            if i < len(extra):
                                user_col = extra[i]
                                lbl = tk.Entry(pair_row, font=("Consolas", 9), fg="#DC2626",
                                               bg="#FECACA", readonlybackground="#FECACA",
                                               relief="flat", bd=0, width=26)
                                lbl.insert(0, user_col)
                                lbl.config(state="readonly")
                                lbl.pack(side=tk.LEFT)
                            else:
                                tk.Label(pair_row, text="(column missing)", font=("Segoe UI", 9, "italic"),
                                         bg="#FEF2F2", fg="#9CA3AF", width=24, anchor="w").pack(side=tk.LEFT)
                            
                            # Arrow
                            tk.Label(pair_row, text="  →  ", bg="#FEF2F2", fg="#6B7280",
                                     font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
                            
                            # Right side: what it should be renamed to (copyable)
                            e = tk.Entry(pair_row, font=("Consolas", 10, "bold"), fg="#1E40AF",
                                         bg="#DBEAFE", readonlybackground="#DBEAFE",
                                         relief="flat", bd=0, width=26)
                            e.insert(0, required_name)
                            e.config(state="readonly")
                            e.pack(side=tk.LEFT, padx=(0, 4))
                        
                        # Show any remaining extra columns that don't pair up
                        remaining_extra = extra[len(missing):]
                        if remaining_extra:
                            tk.Label(fix_frame, text=f"\nExtra columns in your file (not used): {', '.join(remaining_extra)}",
                                     font=("Segoe UI", 8), bg="#FEF2F2", fg="#9CA3AF",
                                     wraplength=500, justify="left", anchor="w").pack(fill=tk.X, pady=(4, 0))
                    
                    if correct:
                        # Section 2: What's already fine
                        ok_frame = tk.Frame(parent, bg="#F0FDF4", padx=10, pady=8)
                        ok_frame.pack(fill=tk.X, pady=(0, 4))
                        tk.Label(ok_frame, text="✓  Already correct:",
                                 font=("Segoe UI", 9, "bold"), bg="#F0FDF4", fg="#166534",
                                 anchor="w").pack(fill=tk.X)
                        tk.Label(ok_frame, text=", ".join(correct), font=("Segoe UI", 9),
                                 bg="#F0FDF4", fg="#15803D", wraplength=500, justify="left",
                                 anchor="w").pack(fill=tk.X)
                    
                    return list(fmt_mapping.keys())
                
                copy_headers_list = []
                
                if not best_variants:
                    tk.Label(content, text="No known formats found for this importer.",
                             font=("Segoe UI", 10), bg="white", fg="#991B1B").pack(anchor="w")
                elif best_score == 0:
                    tk.Label(content, text="We couldn't recognize any of your column headers.\nThis may not be the right format — contact support if needed.",
                             font=("Segoe UI", 9), bg="white", fg="#6B7280", justify="left",
                             wraplength=500).pack(anchor="w", pady=(0, 8))
                    tk.Label(content, text="If this is a known format, rename your columns to exactly match:",
                             font=("Segoe UI", 10), bg="white", fg="#1E293B").pack(anchor="w", pady=(0, 4))
                    _, fmt_mapping = best_variants[0]
                    copy_headers_list = _build_single_variant_guidance(content, fmt_mapping)
                elif len(best_variants) == 1:
                    _, fmt_mapping = best_variants[0]
                    copy_headers_list = _build_single_variant_guidance(content, fmt_mapping)
                else:
                    # Tie — show multiple options
                    tk.Label(content, text="Your file matches multiple known formats equally.\nPick the one that fits your data:",
                             font=("Segoe UI", 9), bg="white", fg="#6B7280", wraplength=500,
                             justify="left").pack(anchor="w", pady=(0, 6))
                    
                    opts = ['A', 'B', 'C', 'D']
                    copy_headers_per_option = {}
                    for idx, (fmt_name, fmt_mapping) in enumerate(best_variants):
                        opt_label = opts[idx] if idx < len(opts) else str(idx + 1)
                        
                        targets = [str(v).lower().strip() for v in fmt_mapping.values()]
                        has_coo = 'country of origin' in targets
                        desc = "with Country of Origin" if has_coo else "without Country of Origin"
                        
                        opt_frame = tk.LabelFrame(content, text=f"  Option {opt_label} ({desc})  ",
                                                  font=("Segoe UI", 9, "bold"), bg="white",
                                                  fg="#1E293B", padx=8, pady=6)
                        opt_frame.pack(fill=tk.X, pady=4)
                        
                        hdrs = _build_single_variant_guidance(opt_frame, fmt_mapping)
                        copy_headers_per_option[opt_label] = hdrs
                    
                    if copy_headers_per_option:
                        copy_headers_list = list(copy_headers_per_option.values())[0]
                
                # Separator + action buttons
                sep = tk.Frame(top, bg="#E5E7EB", height=1)
                sep.pack(fill=tk.X, padx=16)
                
                btn_frame = tk.Frame(top, bg="white", pady=10, padx=16)
                btn_frame.pack(fill=tk.X)
                
                def _copy_headers():
                    if copy_headers_list:
                        tab_separated = "\t".join(copy_headers_list)
                        top.clipboard_clear()
                        top.clipboard_append(tab_separated)
                        copy_btn.config(text="✓ Copied to clipboard!", fg="white", bg="#059669")
                        top.after(1500, lambda: copy_btn.config(
                            text="📋  Copy full header row (paste into Excel Row 1)",
                            fg="white", bg="#2563EB"))
                
                copy_btn = tk.Button(btn_frame,
                                     text="📋  Copy full header row (paste into Excel Row 1)",
                                     command=_copy_headers, font=("Segoe UI", 9, "bold"),
                                     bg="#2563EB", fg="white", relief="flat", padx=12, pady=4,
                                     cursor="hand2")
                copy_btn.pack(side=tk.LEFT)
                
                tk.Button(btn_frame, text="Close", command=top.destroy, font=("Segoe UI", 9),
                          relief="flat", padx=12, pady=4, bg="#F3F4F6", fg="#374151",
                          cursor="hand2").pack(side=tk.RIGHT)
                
                # Size the window to fit content, capped at a reasonable height
                top.update_idletasks()
                req_h = min(content.winfo_reqheight() + 140, 600)
                top.geometry(f"600x{req_h}")
                
                self.root.wait_window(top)
                return

            mapped_cols = {}
            expected_headers_lower = {k.lower(): v for k, v in expected_headers.items()}
            for col in invoice_df.columns:
                col_str = str(col).strip()
                if col_str in expected_headers:
                    mapped_cols[col] = expected_headers[col_str]
                elif col_str.lower() in expected_headers_lower:
                    mapped_cols[col] = expected_headers_lower[col_str.lower()]
                    
            if not mapped_cols:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", "No valid mappings applied.")
                return
                
            model_excel_col = None
            for col, target in mapped_cols.items():
                if target == 'Model' or target.lower() == 'model':
                    model_excel_col = col
                    break
                    
            if not model_excel_col:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showerror("Error", "No mapping found for 'Model'. It is required.")
                return
                
            target_to_col_idx = {}
            for target in set(mapped_cols.values()):
                t_lower = target.lower()
                for i, h in enumerate(checklist_logic.LOGISYS_HEADERS):
                    if str(h).strip().lower() == t_lower:
                        target_to_col_idx[target] = f"col_{i+1}"
                        break
                        
            queue_items = []
            for _, row in invoice_df.iterrows():
                model_val = str(row[model_excel_col]).strip() if pd.notna(row[model_excel_col]) else ""
                if not model_val:
                    continue
                    
                prefilled = {}
                for col, target in mapped_cols.items():
                    if target in target_to_col_idx:
                        val = row[col]
                        # Handle datetime objects
                        if hasattr(val, 'strftime'):
                            val = val.strftime('%d-%m-%Y')
                        prefilled[target_to_col_idx[target]] = str(val).strip() if pd.notna(val) else ""
                        
                queue_items.append({'model': model_val, 'prefilled': prefilled})
                
            if queue_items:
                if not hasattr(self, 'session_queue'):
                    self.session_queue = []
                self.session_queue.extend(queue_items)
                
                self.status_label.config(text=f"⚙️ Matching {len(queue_items)} models...")
                self.root.update_idletasks()
                self._process_model_queue(queue_items)
            else:
                self.root.config(cursor="")
                self.status_label.config(text="")
                messagebox.showinfo("Info", "No valid rows found in the uploaded file.")
                
        _brand_button(input_row, "+ Add Models", do_add_models).pack(side=tk.LEFT, padx=(10,0))
        _brand_button(status_row, "✖ Clear", _do_clear, secondary=True).pack(side=tk.LEFT, padx=(10,0))
        _brand_button(input_row, "📁 Upload Invoice Excel", _do_upload_invoice, secondary=True).pack(side=tk.LEFT, padx=(10,0))
        self.btn_export = _brand_button(input_row, "📥 Generate CSV", self._do_export)
        self.btn_export.pack(side=tk.LEFT, padx=(10,0))
        self.lbl_unresolved_count = tk.Label(input_row, text="", font=("Segoe UI", 9, "bold"), fg="#D97706", bg=_PANEL_WHITE)
        self.lbl_unresolved_count.pack(side=tk.LEFT, padx=(10,0))
        self.status_label.pack(side=tk.RIGHT)
        
        # 3. Tabs Area (Notebook)
        notebook_container = tk.Frame(self.body_container)
        notebook_container.pack(fill=tk.BOTH, expand=True)
        self.notebook = ttk.Notebook(notebook_container)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        tab_preview = tk.Frame(self.notebook, bg=_PANEL_WHITE)
        tab_missing = tk.Frame(self.notebook, bg=_PANEL_WHITE)
        self.notebook.add(tab_preview, text="Checklist Preview")
        self.notebook.add(tab_missing, text="Models Not Found (0)")
        
        # --- Checklist Preview Tab ---
        # Resolution panel (shown only when unresolved multi-match models exist)
        self.resolution_panel = tk.Frame(tab_preview, bg="#FFFBEB", bd=1, relief=tk.SOLID)
        # Packed/unpacked dynamically in _rebuild_resolution_panel
        
        table_frame = tk.Frame(tab_preview, bg=_PANEL_WHITE)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Instantiate tksheet Sheet widget
        wrapped_headers = [wrap_header(h) for h in checklist_logic.LOGISYS_HEADERS]
        if self.importer and self.importer.strip().lower() == 'ansell':
            wrapped_headers[3] = wrap_header("Product Desc [Auto]")
        self.sheet = Sheet(
            table_frame,
            headers=wrapped_headers,
            empty_horizontal=0,
            empty_vertical=0,
            show_row_index=True,
            show_column_header=True,
            row_height=28
        )
        self.sheet.pack(fill=tk.BOTH, expand=True)
        
        # Apply themes and layout options
        self.sheet.set_options(
            table_bg="#F9FAFB",
            even_rows_bg="#F9FAFB",
            odd_rows_bg="#FFFFFF",
            header_bg="#F3F4F6",
            header_fg=_PRIMARY_BLUE,
            header_font=("Segoe UI", 10, "bold"),
            header_grid_color=_BORDER_GRAY,
            grid_color=_BORDER_GRAY,
            selected_cells_bg=_HOVER_BLUE,
            selected_cells_fg="#FFFFFF",
            selected_rows_bg=_HOVER_BLUE,
            selected_rows_fg="#FFFFFF",
            row_index_bg="#F3F4F6",
            row_index_fg=_MUTED_GRAY
        )
        self.sheet.set_header_height_lines(1)
        self.sheet.set_column_widths(GRID_WIDTHS)
        
        # Highlight headers for master columns
        for col in range(len(checklist_logic.LOGISYS_HEADERS)):
            if (col + 1) in MASTER_INDICES:
                self.sheet.highlight_cells(row="all", column=col, canvas="header", bg="#D6E4F0", fg=_PRIMARY_BLUE)
            
        # Enable basic bindings
        self.sheet.enable_bindings(
            "edit_cell",
            "single_select",
            "row_select",
            "column_width_resize",
            "row_height_resize",
            "arrowkeys",
            "copy",
            "rc_select"
        )
        self.sheet.extra_bindings("end_edit_cell", self._on_cell_edited)
        self.root.bind("<Delete>", self._on_delete_key)
        self.root.bind("<BackSpace>", self._on_delete_key)
        self.sheet.popup_menu_add_command("✅ Resolve Conflict: Keep this row", self._resolve_conflict)
        
        # --- Missing Models Tab ---
        session = auth.get_current_session()
        can_add = session.get('Can_Add_Edit_Model', False) if isinstance(session, dict) else False
        
        self.missing_table_frame = tk.Frame(tab_missing, bg=_PANEL_WHITE)
        self.missing_table_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Success Frame overlay when empty
        self.success_frame = tk.Frame(self.missing_table_frame, bg=_PANEL_WHITE)
        self.success_msg = tk.Label(self.success_frame, text="", font=("Segoe UI", 16, "bold"), fg="#10B981", bg=_PANEL_WHITE)
        self.success_msg.pack(expand=True, pady=100)
        
        self.missing_sheet = Sheet(
            self.missing_table_frame,
            headers=[wrap_header(h) for h in PENDING_HEADERS],
            empty_horizontal=0,
            empty_vertical=0,
            show_row_index=True,
            show_column_header=True,
            row_height=28
        )
        self.missing_sheet.set_header_height_lines(1)
        
        # Determine column widths for missing sheet
        missing_widths = []
        for h in PENDING_HEADERS:
            h_clean = h.strip()
            # Ensure the width is at least enough to fit the header text clearly
            min_header_width = len(h_clean) * 7 + 30
            
            if h_clean in NARROW_COLS:
                base_w = 80
            elif h_clean in MEDIUM_COLS:
                base_w = 140
            elif h_clean in WIDE_COLS:
                base_w = 250
            else:
                h_lower = h_clean.lower()
                if any(term in h_lower for term in ['rate', 'sno', 'sr.no', 'srno', 'cth']):
                    base_w = 80
                elif any(term in h_lower for term in ['notn', 'notification']):
                    base_w = 140
                else:
                    base_w = 120
                    
            missing_widths.append(max(base_w, min_header_width))
            
        self.missing_sheet.set_column_widths(missing_widths)
        
        # Hide Importer, Added_By, and Added_At columns from the UI
        cols_to_hide = []
        for col_name in ["Importer", "Added_By", "Added_At"]:
            if col_name in PENDING_HEADERS:
                cols_to_hide.append(PENDING_HEADERS.index(col_name))
        if cols_to_hide:
            self.missing_sheet.hide_columns(columns=cols_to_hide)
        
        # Enable editing globally. We enforce permissions in the event handlers.
        self.missing_sheet.enable_bindings(
            "edit_cell",
            "single_select",
            "drag_select",
            "select_all",
            "column_select",
            "row_select",
            "column_width_resize",
            "row_height_resize",
            "arrowkeys",
            "copy",
            "cut",
            "paste",
            "delete",
            "undo",
            "rc_select"
        )
        self.missing_sheet.extra_bindings("end_edit_cell", self._on_missing_cell_edited)
        self.missing_sheet.extra_bindings("begin_edit_cell", self._on_missing_begin_edit)
        self.missing_sheet.extra_bindings("end_paste", self._on_missing_paste)
        self.missing_sheet.extra_bindings("rc_select", self._on_missing_rc_select)
            
        # Context columns are always read-only: Importer, Added_By, Added_At, Model, Country of Origin
        context_cols = ["Importer", "Added_By", "Added_At", "Model", "Country of Origin"]
        readonly_indices = [PENDING_HEADERS.index(c) for c in context_cols if c in PENDING_HEADERS]
        self.missing_sheet.readonly_columns(columns=readonly_indices, readonly=True)
            
        # Tab Actions Frame (aligned with the Notebook tabs on the top right)
        self.tab_actions_frame = tk.Frame(notebook_container, bg=_LIGHT_BG)
        self.tab_actions_frame.place(relx=1.0, x=-10, y=2, anchor="ne")
        
        self.btn_recheck = _brand_button(self.tab_actions_frame, "🔄 Re-check Models", self._do_recheck_models, secondary=True)
        self.btn_download_missing = _brand_button(self.tab_actions_frame, "📊 Export Excel", self._do_export_missing_excel, secondary=True)
        self.btn_submit_master = _brand_button(self.tab_actions_frame, "📤 Update Master Sheet", self._do_submit_master, bg_color=_PRIMARY_BLUE)
        self.btn_send_review = _brand_button(self.tab_actions_frame, "📤 Send for Review", self._do_send_for_review, bg_color=_PRIMARY_BLUE)
        
        can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
        
        if not can_add and not can_approve:
            self.btn_submit_master.config(state="disabled", bg="#9CA3AF")
            self.btn_send_review.config(state="disabled", bg="#9CA3AF")
            self.view_only_label = tk.Label(self.tab_actions_frame, text="(View Only - No Permission)", fg="#DC2626", bg=_LIGHT_BG, font=("Segoe UI", 9, "bold"))
        else:
            self.view_only_label = None

        def on_tab_changed(event):
            selected_tab = event.widget.select()
            tab_text = event.widget.tab(selected_tab, "text")
            
            # Hide all
            self.btn_recheck.pack_forget()
            self.btn_download_missing.pack_forget()
            self.btn_submit_master.pack_forget()
            self.btn_send_review.pack_forget()
            if self.view_only_label:
                self.view_only_label.pack_forget()
                
            if "Checklist Preview" in tab_text:
                self.btn_recheck.pack(side=tk.RIGHT)
            else:
                session = auth.get_current_session()
                can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
                can_add = session.get('Can_Add_Edit_Model', False) if isinstance(session, dict) else False
                
                if can_approve:
                    self.btn_submit_master.pack(side=tk.RIGHT)
                elif can_add:
                    self.btn_send_review.pack(side=tk.RIGHT)
                    
                self.btn_download_missing.pack(side=tk.RIGHT, padx=(0, 5))
                if self.view_only_label:
                    self.view_only_label.pack(side=tk.RIGHT, padx=(0, 10))
                    
        self.notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

        # Reload current rows (when switching importer screens)
        self._reload_table_rows()
        self._refresh_pending_tab()


    def _reload_table_rows(self):
        self.sheet.pack(fill=tk.BOTH, expand=True)
        
        # Map self.rows_data dict list to data list of lists
        data_list = []
        for row in self.rows_data:
            row_list = []
            for c_idx in range(1, len(checklist_logic.LOGISYS_HEADERS) + 1):
                row_list.append(row.get(f"col_{c_idx}", ""))
            data_list.append(row_list)
            
        self.sheet.set_sheet_data(data_list, redraw=False)
        
        # Apply Nagarkot standard styling options to sheet
        self.sheet.set_options(
            table_bg="#F9FAFB",
            even_rows_bg="#F9FAFB",
            odd_rows_bg="#FFFFFF",
            header_bg="#F3F4F6",
            header_fg=_PRIMARY_BLUE,
            header_font=("Segoe UI", 10, "bold"),
            header_grid_color=_BORDER_GRAY,
            grid_color=_BORDER_GRAY,
            selected_cells_bg=_HOVER_BLUE,
            selected_cells_fg="#FFFFFF",
            selected_rows_bg=_HOVER_BLUE,
            selected_rows_fg="#FFFFFF",
            row_index_bg="#F3F4F6",
            row_index_fg=_MUTED_GRAY
        )
        self.sheet.set_header_height_lines(2)
        self.sheet.set_column_widths(GRID_WIDTHS)
        
        # Dynamically align columns based on data types
        align_sheet_columns(self.sheet, checklist_logic.LOGISYS_HEADERS)
        
        # Clear existing highlights first
        self.sheet.dehighlight_all()
        
        # Highlight headers for master columns
        for col in range(len(checklist_logic.LOGISYS_HEADERS)):
            if (col + 1) in MASTER_INDICES:
                self.sheet.highlight_cells(row="all", column=col, canvas="header", bg="#D6E4F0")
                
        # Highlight missing and unresolved rows cell-by-cell (overriding alternates)
        for r_idx, row in enumerate(self.rows_data):
            if row.get('_missing_master'):
                for col_idx in range(len(checklist_logic.LOGISYS_HEADERS)):
                    self.sheet.highlight_cells(row=r_idx, column=col_idx, bg="#FCA5A5") # light red
            elif row.get('_unresolved_group') or row.get('_awaiting_resolution'):
                for col_idx in range(len(checklist_logic.LOGISYS_HEADERS)):
                    self.sheet.highlight_cells(row=r_idx, column=col_idx, bg="#FDE047") # light yellow
            
        self.sheet.redraw()
        self._rebuild_resolution_panel()
        self._update_missing_tab_visibility()

    def _rebuild_resolution_panel(self):
        # Clear existing widgets
        for widget in self.resolution_panel.winfo_children():
            widget.destroy()
            
        # Find first unresolved group
        target_group = None
        target_model = None
        for row in self.rows_data:
            if row.get('_unresolved_group'):
                target_group = row['_unresolved_group']
                target_model = row.get('col_24', 'Unknown Model')
                break
                
        if not target_group:
            self.resolution_panel.pack_forget()
            return
            
        # We have an unresolved group. Build the UI.
        self.resolution_panel.pack(fill=tk.X, padx=5, pady=5)
        
        header = tk.Label(self.resolution_panel, 
                          text=f"⚠ Model '{target_model}' has multiple master matches. Choose the correct one below:",
                          font=("Segoe UI", 10, "bold"), fg="#D97706", bg="#FFFBEB")
        header.pack(anchor="w", padx=10, pady=(5, 5))
        
        btn_container = tk.Frame(self.resolution_panel, bg="#FFFBEB")
        btn_container.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # Build a button for each candidate row in this group
        for r_idx, row in enumerate(self.rows_data):
            if row.get('_unresolved_group') == target_group:
                cth = row.get('col_9', '')
                notn = row.get('col_10', '')
                srno = row.get('col_11', '')
                details = f"Row {r_idx+1}: CTH {cth}" if cth else f"Row {r_idx+1}: No CTH"
                if notn:
                    details += f" | Notn {notn}"
                    if srno:
                        details += f" (SrNo {srno})"
                        
                row_frame = tk.Frame(btn_container, bg="#FFFBEB")
                row_frame.pack(fill=tk.X, pady=2)
                
                # We need to capture r_idx properly in the lambda
                btn = tk.Button(row_frame, text="✓ Use This Row", bg=_PRIMARY_BLUE, fg="white",
                                font=("Segoe UI", 9, "bold"), relief=tk.FLAT, padx=10, pady=2,
                                command=lambda r=r_idx: self._resolve_conflict_for_row(r))
                btn.pack(side=tk.LEFT, padx=(0, 10))
                
                lbl = tk.Label(row_frame, text=details, font=("Segoe UI", 9), bg="#FFFBEB", fg="#1F2937")
                lbl.pack(side=tk.LEFT)


    def _update_missing_tab_visibility(self):
        try:
            missing_rows = self.missing_sheet.get_sheet_data()
            missing_count = len([r for r in missing_rows if any(r)])
        except Exception:
            missing_count = 0
            
        self.notebook.tab(1, text=f"Models Not Found ({missing_count})")
        
        export_blocked = False
        unresolved_models = set(str(row.get('col_24', '')).strip().lower() for row in self.rows_data if row.get('_unresolved_group'))
        unresolved_count = len(unresolved_models)
        
        missing_models_in_master = set(str(row.get('col_24', '')).strip().lower() for row in self.rows_data if row.get('_missing_master'))
        actual_missing_count = len(missing_models_in_master)
        
        if hasattr(self, 'lbl_unresolved_count'):
            if unresolved_count > 0:
                self.lbl_unresolved_count.config(text=f"⚠ {unresolved_count} model(s) need conflict resolution")
            elif actual_missing_count > 0:
                self.lbl_unresolved_count.config(text=f"⚠ {actual_missing_count} model(s) not in master sheet (export disabled)")
            else:
                self.lbl_unresolved_count.config(text="")
                
        if unresolved_count > 0 or actual_missing_count > 0:
            export_blocked = True
        
        if missing_count == 0:
            if hasattr(self, 'btn_export'):
                self.btn_export.config(state="normal" if not export_blocked else "disabled", 
                                       bg=_PRIMARY_BLUE if not export_blocked else "#9CA3AF")
            self.missing_sheet.pack_forget()
            if hasattr(self, 'rows_data') and self.rows_data:
                self.success_msg.config(text="🎉 No pending models for this importer!")
            else:
                self.success_msg.config(text="")
            self.success_frame.pack(fill=tk.BOTH, expand=True)
        else:
            if hasattr(self, 'btn_export'):
                if export_blocked:
                    self.btn_export.config(state="disabled", bg="#9CA3AF")
                else:
                    self.btn_export.config(state="normal", bg=_PRIMARY_BLUE)
            self.success_frame.pack_forget()
            self.missing_sheet.pack(fill=tk.BOTH, expand=True)
            
            # Apply standard styling and alignments to missing sheet
            self.missing_sheet.set_options(
                table_bg="#F9FAFB",
                even_rows_bg="#F9FAFB",
                odd_rows_bg="#FFFFFF",
                header_bg="#F3F4F6",
                header_fg=_PRIMARY_BLUE,
                header_font=("Segoe UI", 10, "bold"),
                header_grid_color=_BORDER_GRAY,
                grid_color=_BORDER_GRAY,
                selected_cells_bg=_HOVER_BLUE,
                selected_cells_fg="#FFFFFF",
                row_index_bg="#F3F4F6",
                row_index_fg=_MUTED_GRAY
            )
            align_sheet_columns(self.missing_sheet, PENDING_HEADERS)
            self.missing_sheet.redraw()


    def _on_cell_edited(self, event):
        r = event.row
        c = event.column
        new_val = str(event.value).strip()
        col_key = f"col_{c+1}"
        
        # Validate cell input
        success, err_msg = self._validate_cell(f"row_{r}", col_key, new_val)
        if not success:
            messagebox.showerror("Validation Error", err_msg)
            # Revert cell value in sheet
            old_val = self.rows_data[r].get(col_key, '')
            self.sheet.set_cell_data(r, c, old_val, redraw=True)
            return
            
        self.rows_data[r][col_key] = new_val

    def _resolve_conflict(self, event=None):
        try:
            cells = self.sheet.get_selected_cells()
        except Exception:
            return
        if not cells: return
        r = cells[0][0]
        self._resolve_conflict_for_row(r)
        
    def _resolve_conflict_for_row(self, r):
        if r >= len(self.rows_data): return
        
        row_data = self.rows_data[r]
        group_id = row_data.get('_unresolved_group')
        if not group_id: return
        
        model = row_data.get('col_24')
        model_key = "".join(str(model).split()).lower()
        
        cth = row_data.get('col_9', '')
        notn = row_data.get('col_10', '')
        srno = row_data.get('col_11', '')
        rate = row_data.get('col_15', '') # In case there's a rate, though not mapped by default
        details = f"CTH: {cth}" if cth else "No CTH"
        if notn:
            details += f" | Basic Notn: {notn}"
            if srno:
                details += f" (SrNo: {srno})"
                
        # Count total occurrences (1 group of candidates + N placeholders)
        count_subsequent = sum(1 for r_row in self.rows_data if r_row.get('_awaiting_resolution') == model_key)
        total_occurrences = 1 + count_subsequent
        
        msg = f"You selected:\n\nModel: {model}\n{details}\n\nApply this master data to all {total_occurrences} occurrence(s) of this model in the checklist?"
        if not messagebox.askyesno("Confirm Resolution", msg):
            return
            
        master_keys = ['col_4', 'col_9', 'col_10', 'col_11', 'col_22', 'col_25', 'col_32', 'col_41', 'col_42', 'col_54', 'col_55', 'col_81', 'col_82']
        
        new_rows = []
        for r_idx, row in enumerate(self.rows_data):
            if row.get('_unresolved_group') == group_id:
                if r_idx != r:
                    continue # Skip this un-picked candidate
                else:
                    # Remove the unresolved flag from the picked candidate
                    row.pop('_unresolved_group', None)
                    new_rows.append(row)
            elif row.get('_awaiting_resolution') == model_key:
                # Update placeholder row with master data from picked candidate
                for mk in master_keys:
                    row[mk] = row_data.get(mk, "")
                row.pop('_awaiting_resolution', None)
                new_rows.append(row)
            else:
                new_rows.append(row)
            
        self.rows_data = new_rows
        self._reload_table_rows()

    def _on_delete_key(self, event):
        if isinstance(event.widget, (tk.Entry, tk.Text)):
            return
        if not hasattr(self, 'sheet') or not self.sheet.winfo_viewable():
            return
        try:
            cells = self.sheet.get_selected_cells()
        except Exception:
            return
        if not cells:
            return
        modified = False
        for r, c in cells:
            col_key = f"col_{c+1}"
            if self.rows_data[r].get(col_key, "") != "":
                self.rows_data[r][col_key] = ""
                self.sheet.set_cell_data(r, c, "", redraw=False)
                modified = True
        if modified:
            self.sheet.redraw()

    def _process_model_queue(self, queue):
        import uuid
        
        models_to_push = []
        # Track seen multi-match models: model_str -> group_id
        seen_multi_match = {}
        
        for item in queue:
            if isinstance(item, dict):
                model = item.get('model', '')
                prefilled = item.get('prefilled', {})
            else:
                model = item
                prefilled = {}
                
            model_key = "".join(str(model).split()).lower()
                
            matches = checklist_logic.lookup_model_local(self.importer_df, model)
            
            if len(matches) == 0:
                self._add_row_to_table(model, {}, prefilled, _missing_master=True)
                models_to_push.append({'model': model, 'prefilled': prefilled})
                
            elif len(matches) == 1:
                self._add_row_to_table(model, matches[0], prefilled)
            else:
                if model_key in seen_multi_match:
                    # Already generated candidates for this model. Add a placeholder row.
                    self._add_row_to_table(model, matches[0], prefilled, _awaiting_resolution=model_key)
                else:
                    # Add all candidate rows with a unique group ID
                    group_id = str(uuid.uuid4())
                    seen_multi_match[model_key] = group_id
                    for m in matches:
                        self._add_row_to_table(model, m, prefilled, _unresolved_group=group_id)
                    
        if models_to_push:
            # Send to Pending_Model_Approval and then refresh tab
            def _push_pending():
                try:
                    from gsheets import GSheetsClient
                    client = GSheetsClient(self.cred_path.get())
                    url = self.gsheets_url.get()
                    session = auth.get_current_session()
                    username = session.get('Username', 'User') if isinstance(session, dict) else 'User'
                    
                    pending_rows = []
                    for m_info in models_to_push:
                        model = m_info['model']
                        prefilled = m_info['prefilled']
                        
                        # Extract any mapped fields from invoice excel
                        extracted_vals = {"Product Desc": "", "Country of Origin": "", "Generic Description": ""}
                        for h_idx, header_name in enumerate(checklist_logic.LOGISYS_HEADERS):
                            col_key = f"col_{h_idx+1}"
                            if col_key in prefilled:
                                clean_header = str(header_name).strip().lower()
                                for k in extracted_vals.keys():
                                    if k.lower() == clean_header:
                                        extracted_vals[k] = prefilled[col_key]
                                        
                        if self.importer and self.importer.strip().lower() == 'ansell' and '_ansell_raw_desc' in prefilled:
                            raw_desc = prefilled.get('_ansell_raw_desc', '')
                            ship_case = str(prefilled.get('_ansell_case', '')).strip()
                            if ship_case.endswith('.00'):
                                ship_case = ship_case[:-3]
                            curr = prefilled.get('col_7') or extracted_vals.get('Currency', '')
                            unit_price = prefilled.get('_ansell_unit_price', '')
                            # Empty Generic Description for pending queue
                            extracted_vals["Product Desc"] = f"PRODUCT CODE. {model} {raw_desc} () [QTY {ship_case} CTN @ {curr} {unit_price} PER CTN]"
                                        
                        pr = {
                            "Model": model,
                            "Product Desc": extracted_vals["Product Desc"],
                            "Country of Origin": extracted_vals["Country of Origin"],
                            "Generic Description": extracted_vals["Generic Description"],
                            "Importer": self.importer,
                            "Added_By": username,
                            "Added_At": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            "Status": "Draft"
                        }
                        pending_rows.append(pr)
                        
                    client.add_pending_models(url, pending_rows)
                except Exception as e:
                    print(f"Failed to push to Pending Model Approval queue: {e}")
                finally:
                    # Refresh the tab to show newly pending items
                    self.root.after(0, self._refresh_pending_tab)
                    
            threading.Thread(target=_push_pending, daemon=True).start()
        else:
            self._update_missing_tab_visibility()

        self.root.config(cursor="")
        if hasattr(self, 'status_label'):
            self.status_label.config(text="")

    def _add_row_to_table(self, model, master_data, prefilled=None, _missing_master=False, _unresolved_group=None, _awaiting_resolution=None):
        prefilled = prefilled or {}
        # Auto-inherit invoice header details from existing rows if typed
        inv_no, inv_date, toi, currency, brand = "", "", "", "", ""
        if self.rows_data and not prefilled:
            first_row = self.rows_data[0]
            inv_no = first_row.get('col_1', '')
            inv_date = first_row.get('col_2', '')
            toi = first_row.get('col_3', '')
            currency = first_row.get('col_7', '')
            brand = first_row.get('col_23', '')
            
        row_data = {}
        for c_idx in range(1, len(checklist_logic.LOGISYS_HEADERS) + 1):
            row_data[f"col_{c_idx}"] = ""
            
        for col_key, val in prefilled.items():
            row_data[col_key] = val
            
        # Set invoice-level columns if not already populated
        if not row_data.get('col_1'): row_data['col_1'] = inv_no
        if not row_data.get('col_2'): row_data['col_2'] = inv_date
        if not row_data.get('col_3'): row_data['col_3'] = toi
        if not row_data.get('col_7'): row_data['col_7'] = currency
        if not row_data.get('col_23'): row_data['col_23'] = brand
        
        # Map master-sourced columns using case-flexible matching (ONLY if not mapped from invoice excel already)
        row_data['col_24'] = model # Model
        
        # get_flexible is defined at module level in this file
        
        # Country of Origin Logic: Exclusively from invoice, UNLESS active variant has no COO mapping
        has_coo_mapping = False
        if hasattr(self, 'active_variant_targets'):
            for target in self.active_variant_targets:
                if str(target).strip().lower() == 'country of origin':
                    has_coo_mapping = True
                    break
                    
        if not has_coo_mapping and 'col_32' not in prefilled:
            row_data['col_32'] = ""
            
        if self.importer and self.importer.strip().upper() in ('BBRAUN', 'AVIAT'):
            row_data['col_12'] = "NOEXCISE"
            row_data['col_39'] = "N"
            
        def safe_set(col_key, md_key):
            if not row_data.get(col_key) and master_data:
                row_data[col_key] = get_flexible(master_data, md_key)
                
        if self.importer and self.importer.strip().lower() == 'ansell':
            # Dynamic construction for Ansell (even if missing master)
            raw_desc = prefilled.get('_ansell_raw_desc', '')
            gen_desc = get_flexible(master_data, 'Generic Description') if master_data else ''
            ship_case = str(prefilled.get('_ansell_case', '')).strip()
            if ship_case.endswith('.00'):
                ship_case = ship_case[:-3]
            curr = prefilled.get('col_7') or row_data.get('col_7', '')
            unit_price = prefilled.get('_ansell_unit_price', '')
            
            # PRODUCT CODE. {Model} {raw} ({Generic}) [QTY {Ship Case} CTN @ {Currency} {Unit Price} PER CTN]
            # Enforce exactly one space after "PRODUCT CODE."
            constructed = f"PRODUCT CODE. {model} {raw_desc} ({gen_desc}) [QTY {ship_case} CTN @ {curr} {unit_price} PER CTN]"
            row_data['col_4'] = constructed
                
            safe_set('col_9', 'CTH')
            if not row_data.get('col_9'):
                row_data['col_9'] = prefilled.get('_ansell_hs', '')
        else:
            if master_data:
                row_data['col_4'] = get_flexible(master_data, 'Product Desc')
            else:
                safe_set('col_4', 'Product Desc')
            safe_set('col_9', 'CTH')
        safe_set('col_10', 'Basic Duty Notn')
        safe_set('col_11', 'Basic Duty Notn SNo')
        safe_set('col_22', 'Generic Description')
        safe_set('col_25', 'End Use')
        safe_set('col_41', 'IGST Notification')
        safe_set('col_42', 'IGST Notification SrNo')
        safe_set('col_54', 'SWS Notification')
        safe_set('col_55', 'SWS Notification SrNo')
        safe_set('col_81', 'AIDC Notn (Customs)')
        safe_set('col_82', 'AIDC Notn Sr.No.(Customs)')
        
        # Set manual default values
        

        if _missing_master: row_data['_missing_master'] = True
        if _unresolved_group: row_data['_unresolved_group'] = _unresolved_group
        if _awaiting_resolution: row_data['_awaiting_resolution'] = _awaiting_resolution
        
        self.rows_data.append(row_data)
        
        # Reload spreadsheet to reflect additions
        self._reload_table_rows()

    def _validate_cell(self, row_id, col_name, new_val):
        if col_name == 'col_5': # Quantity
            if not new_val:
                return False, "'Quantity' cannot be empty."
            try:
                float(new_val)
            except ValueError:
                return False, "'Quantity' must be numeric."
        elif col_name in ['col_8', 'col_16']: # Rate or Amount
            if new_val:
                try:
                    float(new_val)
                except ValueError:
                    name = "Rate" if col_name == 'col_8' else "Amount"
                    return False, f"'{name}' must be numeric."
        return True, ""

    def _do_export(self):
        if not self.rows_data:
            messagebox.showerror("Error", "No rows added to the checklist.")
            return
            
        if not (self.importer and self.importer.strip().lower() == 'aviat'):
            for row in self.rows_data:
                if row.get('_unresolved_group') or row.get('_awaiting_resolution'):
                    messagebox.showerror("Validation Error", "You have unresolved model conflicts (highlighted in yellow).\nRight click the correct candidate row and select '✅ Resolve Conflict: Keep this row'.")
                    return
                
        missing_models = set(str(row.get('col_24', '')).strip() for row in self.rows_data if row.get('_missing_master'))
        n_missing = len(missing_models)
        if n_missing > 0:
            messagebox.showerror(
                "Missing Models",
                f"{n_missing} model(s) are not present in the master sheet. Generate CSV is disabled until all models are properly written to the master sheet."
            )
            return
            
        default_name = f"checklist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        p = filedialog.asksaveasfilename(
            defaultextension=".csv", 
            filetypes=[("CSV Files", "*.csv")], 
            title="Save Checklist",
            initialfile=default_name
        )
        if not p:
            return
            
        try:
            checklist_logic.export_checklist(self.rows_data, p)
            
            # Log audit trail
            models_added = [r['col_24'] for r in self.rows_data]
            checklist_logic.log_audit_action(self.gsheets_url.get(), self.cred_path.get(), auth.get_current_session(), self.importer, models_added)
            
            messagebox.showinfo("Success", f"Checklist successfully generated!\nSaved to: {p}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export template: {str(e)}")

    def _do_export_missing_excel(self):
        missing_rows = self.missing_sheet.get_sheet_data()
        valid_rows = [r for r in missing_rows if any(str(x).strip() for x in r)]
        if not valid_rows:
            messagebox.showinfo("Info", "No missing models available to export.")
            return
            
        default_name = f"missing_models_{self.importer or 'export'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Models Not Found Excel",
            initialfile=default_name
        )
        if not p:
            return
            
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Models Not Found"

            # Exclude Added_By, Added_At, and Importer from export
            exclude_cols = {"Added_By", "Added_At", "Importer"}
            export_headers = [h for h in PENDING_HEADERS if h not in exclude_cols]

            # Convert valid_rows into list of dicts for easy column access
            export_row_dicts = []
            for r_data in valid_rows:
                r_dict = {h: r_data[i] if i < len(r_data) else "" for i, h in enumerate(PENDING_HEADERS)}
                export_row_dicts.append(r_dict)

            # Check which candidate columns are all-empty across all rows
            highlight_candidate_fields = {"model", "product desc", "cth", "country of origin", "end use", "generic description"}
            cols_to_highlight = set()
            for header in export_headers:
                h_lower = header.strip().lower()
                if h_lower in highlight_candidate_fields:
                    is_all_empty = all(not str(r_dict.get(header, '')).strip() for r_dict in export_row_dicts)
                    if not is_all_empty:
                        cols_to_highlight.add(h_lower)

            # Required headers to highlight in soft blue
            req_headers = {
                "Model", "Product Desc", "Country of Origin", "CTH", "End Use", "Generic Description",
                "Basic Duty Notn", "Basic Duty Notn SNo", "Basic Duty Rate",
                "Customs Health Cess Notn", "Customs Health Cess SNo", "Customs Health Cess Rate",
                "SWS Notification", "SWS Notification SrNo", "SWS Rate",
                "IGST Notification", "IGST Notification SrNo", "IGST Rate",
                "AIDC Notn (Customs)", "AIDC Notn Sr.No.(Customs)"
            }

            req_header_fill = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")
            default_header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="1F3F6E")
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )

            ws.append(export_headers)
            for col_idx, header in enumerate(export_headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.border = thin_border
                if header.strip() in req_headers:
                    cell.fill = req_header_fill
                else:
                    cell.fill = default_header_fill

            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

            for r_dict in export_row_dicts:
                row_data = [r_dict.get(h, "") for h in export_headers]
                ws.append(row_data)
                current_row = ws.max_row
                for col_idx, header in enumerate(export_headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    val_str = str(cell.value or '').strip()
                    if header.strip().lower() in cols_to_highlight and not val_str:
                        cell.fill = yellow_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(p)
            messagebox.showinfo("Success", f"Models Not Found table exported successfully to Excel:\n{p}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel file: {str(e)}")

    def _do_master_search(self):
        query = self.search_var.get().strip() if hasattr(self, 'search_var') else ""
        if not query:
            messagebox.showinfo("Search Master Sheet", "Please enter a search keyword or CTH in the 'Search Master Sheet:' box above.")
            return

        if not hasattr(self, 'importer_df') or self.importer_df is None or self.importer_df.empty:
            messagebox.showinfo("Search Master Sheet", f"No master sheet data loaded for importer '{self.importer or 'Unknown'}'.")
            return

        query_lower = query.lower()
        df = self.importer_df.copy()

        def matches(row):
            m = str(row.get('Model', '')).lower()
            d = str(row.get('Product Desc', '')).lower()
            c = str(row.get('CTH', '')).lower()
            g = str(row.get('Generic Description', '')).lower()
            return (query_lower in m) or (query_lower in d) or (query_lower in c) or (query_lower in g)

        matched_df = df[df.apply(matches, axis=1)]
        rows_data = []
        for _, row in matched_df.iterrows():
            r_vals = [str(row.get(h, '')).strip() for h in MASTER_SHEET_COLUMNS]
            rows_data.append(r_vals)

        popup = tk.Toplevel(self.root)
        self._search_popup = popup
        popup.title(f"Search Results for '{query}' - {self.importer or 'Master Sheet'}")
        popup.geometry("980x560")
        popup.configure(bg="#F4F6F8")

        # Dark navy header banner matching user screenshot
        header_banner = tk.Frame(popup, bg="#1F3F6E", height=48)
        header_banner.pack(fill=tk.X)
        header_banner.pack_propagate(False)

        lbl_header = tk.Label(
            header_banner,
            text=f"🔍 Found {len(rows_data)} matching rows",
            font=("Segoe UI", 11, "bold"),
            fg="#FFFFFF",
            bg="#1F3F6E"
        )
        lbl_header.pack(side=tk.LEFT, padx=16, pady=10)

        # Table Container
        table_container = tk.Frame(popup, bg=_PANEL_WHITE)
        table_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        popup_sheet = Sheet(
            table_container,
            headers=[wrap_header(h) for h in MASTER_SHEET_COLUMNS],
            empty_horizontal=0,
            empty_vertical=0,
            show_row_index=True,
            show_column_header=True,
            row_height=28
        )
        popup_sheet.set_header_height_lines(1)
        popup_sheet.set_sheet_data(rows_data)
        popup_sheet.pack(fill=tk.BOTH, expand=True)

        popup_sheet.enable_bindings(
            "single_select",
            "drag_select",
            "select_all",
            "column_select",
            "row_select",
            "column_width_resize",
            "row_height_resize",
            "arrowkeys",
            "copy",
            "rc_select"
        )
        popup_sheet.readonly_columns(columns=list(range(len(MASTER_SHEET_COLUMNS))), readonly=True)

        # Bottom Action Bar
        bottom_bar = tk.Frame(popup, bg="#F4F6F8")
        bottom_bar.pack(fill=tk.X, padx=16, pady=(0, 12))

        def apply_selected_row_as_template():
            selected_popup_rows = popup_sheet.get_selected_rows()
            sheet_data = popup_sheet.get_sheet_data()
            if not selected_popup_rows or not sheet_data:
                messagebox.showinfo("Select Row", "Please click/select a row in the search results table first.", parent=popup)
                return

            p_idx = list(selected_popup_rows)[0]
            if p_idx >= len(sheet_data):
                return
            row_vals = sheet_data[p_idx]
            master_rec = {h: row_vals[i] for i, h in enumerate(MASTER_SHEET_COLUMNS)}

            self._apply_template_to_selected_row(master_rec)

        btn_use_template = tk.Button(
            bottom_bar,
            text="📋 Use Selected Row as Template",
            font=("Segoe UI", 9, "bold"),
            fg="white",
            bg="#1F3F6E",
            relief=tk.FLAT,
            cursor="hand2",
            padx=12,
            pady=4,
            command=apply_selected_row_as_template
        )
        btn_use_template.pack(side=tk.LEFT)

        btn_close = tk.Button(
            bottom_bar,
            text="Close",
            font=("Segoe UI", 9),
            fg="#374151",
            bg="#E5E7EB",
            relief=tk.FLAT,
            cursor="hand2",
            padx=16,
            pady=4,
            command=popup.destroy
        )
        btn_close.pack(side=tk.RIGHT)

    def _apply_template_to_selected_row(self, master_record):
        if not hasattr(self, 'missing_sheet') or not self.missing_sheet.winfo_exists():
            return

        selected_rows = self.missing_sheet.get_selected_rows()
        all_sheet_data = self.missing_sheet.get_sheet_data()

        if selected_rows:
            target_indices = sorted(list(selected_rows))
        else:
            valid_indices = [i for i, r in enumerate(all_sheet_data) if any(str(x).strip() for x in r)]
            target_indices = [valid_indices[0]] if valid_indices else []

        if not target_indices:
            messagebox.showinfo("Select Row", "Please click/select a row in the Models Not Found table to apply this template.")
            return

        template_fields = {
            "CTH": master_record.get("CTH", ""),
            "Basic Duty Notn": master_record.get("Basic Duty Notn", ""),
            "Basic Duty Notn SNo": master_record.get("Basic Duty Notn SNo", ""),
            "Basic Duty Rate": master_record.get("Basic Duty Rate", ""),
            "Customs Health Cess Notn": master_record.get("Customs Health Cess Notn", ""),
            "Customs Health Cess SNo": master_record.get("Customs Health Cess SNo", ""),
            "Customs Health Cess Rate": master_record.get("Customs Health Cess Rate", ""),
            "SWS Notification": master_record.get("SWS Notification", ""),
            "SWS Notification SrNo": master_record.get("SWS Notification SrNo", ""),
            "SWS Rate": master_record.get("SWS Rate", ""),
            "IGST Notification": master_record.get("IGST Notification", ""),
            "IGST Notification SrNo": master_record.get("IGST Notification SrNo", ""),
            "IGST Rate": master_record.get("IGST Rate", ""),
            "AIDC Notn (Customs)": master_record.get("AIDC Notn (Customs)", ""),
            "AIDC Notn Sr.No.(Customs)": master_record.get("AIDC Notn Sr.No.(Customs)", ""),
            "End Use": master_record.get("End Use", ""),
            "Generic Description": master_record.get("Generic Description", "")
        }

        model_ref = str(master_record.get("Model", "")).strip()

        for r_idx in target_indices:
            for field_name, val in template_fields.items():
                if field_name in PENDING_HEADERS:
                    c_idx = PENDING_HEADERS.index(field_name)
                    self.missing_sheet.set_cell_data(r_idx, c_idx, str(val or '').strip(), redraw=False)

        self.missing_sheet.redraw()

        row_str = ", ".join([str(i+1) for i in target_indices])
        self.search_status_lbl.config(
            text=f"✓ Applied template '{model_ref}' to row(s) {row_str}.",
            fg="#10B981"
        )

    def _refresh_pending_tab(self):
        if not self.importer:
            return
            
        self.missing_sheet.set_sheet_data([])
        self.btn_submit_master.config(state="disabled")
        self.btn_recheck.config(state="disabled")
        self.root.update_idletasks()
        
        def load_thread():
            try:
                from gsheets import GSheetsClient
                client = GSheetsClient(self.cred_path.get())
                df, _ = client.get_sheet_data(self.gsheets_url.get(), "Pending_Model_Approval")
                
                rows_to_show = []
                if not df.empty:
                    # Filter by relevant statuses and importer
                    valid_statuses = {'draft', 'sent for review', 'need correction', 'approved', 'rejected', 'pending'}
                    pending_df = df[(df['Status'].str.strip().str.lower().isin(valid_statuses)) & (df['Importer'].str.strip() == self.importer)]
                    for _, row in pending_df.iterrows():
                        r_data = []
                        for h in PENDING_HEADERS:
                            r_data.append(str(row.get(h, '')).strip())
                        rows_to_show.append(r_data)
                
                self.root.after(0, lambda d=rows_to_show: self._on_pending_loaded(d))
            except Exception as e:
                self.root.after(0, lambda err=e: self._on_pending_err(err))
                
        threading.Thread(target=load_thread, daemon=True).start()
        
    def _on_pending_loaded(self, data):
        if hasattr(self, 'btn_recheck') and self.btn_recheck.winfo_exists():
            self.btn_recheck.config(state="normal")
            
        can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
        if can_approve and hasattr(self, 'btn_submit_master') and self.btn_submit_master.winfo_exists():
            self.btn_submit_master.config(state="normal", text="📤 Update Master Sheet")
            

                    
        if hasattr(self, 'missing_sheet') and self.missing_sheet.winfo_exists():
            self.missing_sheet.set_sheet_data(data)
            self.missing_sheet.redraw()
            self._update_missing_tab_visibility()
            
    def _on_pending_err(self, err):
        if hasattr(self, 'btn_recheck') and self.btn_recheck.winfo_exists():
            self.btn_recheck.config(state="normal")
        messagebox.showerror("Error", f"Failed to load pending models:\n{err}")

    def _do_submit_master(self):
        selected_rows = self.missing_sheet.get_selected_rows()
        all_sheet_data = self.missing_sheet.get_sheet_data()
        
        # If specific rows are selected, use them; otherwise use all non-empty rows
        if selected_rows:
            target_indices = sorted(list(selected_rows))
        else:
            target_indices = [i for i, r in enumerate(all_sheet_data) if any(str(x).strip() for x in r)]
            
        if not target_indices:
            messagebox.showerror("Error", "No rows available to submit to Master Sheet.")
            return
            
        rows_to_process = []
        for idx in target_indices:
            if idx < len(all_sheet_data):
                r_data = all_sheet_data[idx]
                if any(str(x).strip() for x in r_data):
                    r_dict = {h: r_data[i] if i < len(r_data) else "" for i, h in enumerate(PENDING_HEADERS)}
                    rows_to_process.append(r_dict)
                    
        if not rows_to_process:
            messagebox.showerror("Error", "No valid rows found to submit.")
            return

        # Ansell specific: Auto-fill Generic Description into Product Desc template if present
        for row_dict in rows_to_process:
            if self.importer and self.importer.strip().lower() == 'ansell':
                gen_desc = row_dict.get("Generic Description", "").strip()
                prod_desc = row_dict.get("Product Desc", "")
                if gen_desc and " () " in prod_desc:
                    row_dict["Product Desc"] = prod_desc.replace(" () ", f" ({gen_desc}) ")

        # Validation across all rows
        duty_pairs = [
            ("Basic Duty", "Basic Duty Notn", "Basic Duty Notn SNo", "Basic Duty Rate"),
            ("Customs Health Cess", "Customs Health Cess Notn", "Customs Health Cess SNo", "Customs Health Cess Rate"),
            ("SWS", "SWS Notification", "SWS Notification SrNo", "SWS Rate"),
            ("IGST", "IGST Notification", "IGST Notification SrNo", "IGST Rate"),
            ("AIDC", "AIDC Notn (Customs)", "AIDC Notn Sr.No.(Customs)", None)
        ]

        for row_dict in rows_to_process:
            model_name = row_dict.get("Model", "Unknown")
            if not row_dict.get("CTH"):
                messagebox.showerror("Validation Error", f"Model '{model_name}' has an empty CTH. CTH cannot be empty.")
                return

            for d_name, notn_col, srno_col, rate_col in duty_pairs:
                has_notn = bool(row_dict.get(notn_col))
                has_srno = bool(row_dict.get(srno_col))
                has_rate = bool(row_dict.get(rate_col)) if rate_col else False
                
                if has_notn or has_srno or has_rate:
                    if rate_col:
                        if not ((has_notn and has_srno) ^ has_rate):
                            if not messagebox.askyesno("Duty Validation", f"For model '{model_name}' ({d_name}), you must provide EITHER (Notification + SrNo) OR Rate.\n\nDo you want to proceed anyway?"):
                                return
                    else:
                        if has_notn != has_srno:
                            if not messagebox.askyesno("Duty Validation", f"For model '{model_name}' ({d_name}), you must provide BOTH Notification and SrNo.\n\nDo you want to proceed anyway?"):
                                return

        self.btn_submit_master.config(state="disabled", text=f"⏳ Updating {len(rows_to_process)} model(s)...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        def submit_thread():
            try:
                session = auth.get_current_session()
                if not session or not session.get('Can_Approve_Model', False):
                    raise Exception("You do not have permission to approve and update the Master Sheet.")
                    
                from gsheets import GSheetsClient
                client = GSheetsClient(self.cred_path.get())
                url = self.gsheets_url.get()
                sheet_id = client.extract_sheet_id(url)
                spreadsheet = client.client.open_by_key(sheet_id)

                importer_name = rows_to_process[0].get("Importer") or self.importer
                
                try:
                    worksheet = spreadsheet.worksheet(importer_name)
                except Exception:
                    raise Exception(f"Importer sheet '{importer_name}' not found.")
                    
                live_headers = [str(h) for h in worksheet.row_values(1)]
                if len(live_headers) != len(MASTER_SHEET_COLUMNS):
                    raise Exception(f"Master sheet structure count mismatch for {importer_name}.")
                    
                for pos, (live_h, expected_h) in enumerate(zip(live_headers, MASTER_SHEET_COLUMNS)):
                    if live_h != expected_h:
                        raise Exception(f"Master sheet header mismatch at column {pos + 1}.")
                        
                master_rows_to_append = []
                models_submitted = []
                for row_dict in rows_to_process:
                    master_row = [""] * len(MASTER_SHEET_COLUMNS)
                    for h in MASTER_SHEET_COLUMNS:
                        if h in row_dict:
                            master_row[MASTER_SHEET_COLUMNS.index(h)] = row_dict[h]
                    master_rows_to_append.append(master_row)
                    models_submitted.append(row_dict.get("Model"))

                client.append_rows(worksheet, master_rows_to_append)
                checklist_logic.log_audit_action(url, self.cred_path.get(), session, importer_name, models_submitted, action_type="models_added_bulk")

                for m in models_submitted:
                    client.mark_pending_model_resolved(url, m, importer_name)

                self.root.after(0, lambda count=len(master_rows_to_append): self._on_submit_success(count))
            except Exception as e:
                self.root.after(0, lambda err=e: self._on_submit_error(err))

        threading.Thread(target=submit_thread, daemon=True).start()

    def _on_submit_success(self, count=1):
        if hasattr(self, 'btn_submit_master') and self.btn_submit_master.winfo_exists():
            self.btn_submit_master.config(state="normal", text="📤 Update Master Sheet")
        self.root.config(cursor="")
        messagebox.showinfo("Success", f"Successfully appended {count} model(s) to Master Sheet and marked as resolved.")
        self._refresh_pending_tab()

    def _on_submit_error(self, err):
        if hasattr(self, 'btn_submit_master') and self.btn_submit_master.winfo_exists():
            self.btn_submit_master.config(state="normal", text="📤 Update Master Sheet")
        self.root.config(cursor="")
        messagebox.showerror("Error", f"Failed to submit to master:\n{err}")
        
    def _do_send_for_review(self):
        selected_rows = self.missing_sheet.get_selected_rows()
        all_sheet_data = self.missing_sheet.get_sheet_data()
        
        if selected_rows:
            target_indices = sorted(list(selected_rows))
        else:
            target_indices = [i for i, r in enumerate(all_sheet_data) if any(str(x).strip() for x in r)]
            
        if not target_indices:
            messagebox.showerror("Error", "No rows available to send for review.")
            return
            
        rows_to_process = []
        for idx in target_indices:
            if idx < len(all_sheet_data):
                r_data = all_sheet_data[idx]
                if any(str(x).strip() for x in r_data):
                    r_dict = {h: r_data[i] if i < len(r_data) else "" for i, h in enumerate(PENDING_HEADERS)}
                    status = str(r_dict.get("Status", "")).strip().lower()
                    if status not in ["draft", "need correction", "pending", ""]:
                        messagebox.showerror("Error", f"Model '{r_dict.get('Model')}' cannot be sent for review (Status is '{r_dict.get('Status')}').")
                        return
                    rows_to_process.append(r_dict)
                    
        if not rows_to_process:
            messagebox.showerror("Error", "No valid rows found to send.")
            return
            
        self.btn_send_review.config(state="disabled", text=f"⏳ Sending {len(rows_to_process)} model(s)...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()
        
        def send_thread():
            try:
                from gsheets import GSheetsClient
                client = GSheetsClient(self.cred_path.get())
                url = self.gsheets_url.get()
                
                for row_dict in rows_to_process:
                    client.update_pending_model_row(url, row_dict["Model"], self.importer, {"Status": "Sent for Review"})
                    
                self.root.after(0, lambda: self._on_send_success())
            except Exception as e:
                self.root.after(0, lambda err=e: self._on_send_error(err))
                
        threading.Thread(target=send_thread, daemon=True).start()
        
    def _on_send_success(self):
        if hasattr(self, 'btn_send_review') and self.btn_send_review.winfo_exists():
            self.btn_send_review.config(state="normal", text="📤 Send for Review")
        self.root.config(cursor="")
        messagebox.showinfo("Success", "Successfully sent selected models for review.")
        self._refresh_pending_tab()
        
    def _on_send_error(self, err):
        if hasattr(self, 'btn_send_review') and self.btn_send_review.winfo_exists():
            self.btn_send_review.config(state="normal", text="📤 Send for Review")
        self.root.config(cursor="")
        messagebox.showerror("Error", f"Failed to send for review:\n{err}")
        
    def _on_missing_paste(self, event):
        """Handle paste (Ctrl+V) into the missing models sheet.
        Triggers CTH autofill and auto-save for pasted cells."""
        import traceback as _tb
        def _dbg(msg):
            try:
                with open("autofill_debug.log", "a", encoding="utf-8") as f:
                    f.write(f"{datetime.now().isoformat()} | PASTE: {msg}\n")
            except Exception:
                pass
        
        try:
            _dbg(f"Paste event fired. Event keys: {list(event.keys()) if isinstance(event, dict) else 'N/A'}")
            
            # After paste, re-read the sheet data to find what changed
            all_data = self.missing_sheet.get_sheet_data()
            
            # Check permissions
            session = auth.get_current_session()
            can_add = session.get('Can_Add_Edit_Model', False) if isinstance(session, dict) else False
            can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
            
            if not can_add and not can_approve:
                _dbg("Permission denied")
                return
            
            # Get the cells that were modified by the paste
            # In tksheet 7.x, event.cells.table is a dict of {(row, col): value}
            modified_cells = {}
            if hasattr(event, 'cells') and hasattr(event.cells, 'table'):
                modified_cells = event.cells.table
                _dbg(f"Modified cells from event: {dict(modified_cells)}")
            
            if not modified_cells:
                _dbg("No modified cells found in event, scanning selected cells instead")
                # Fallback: check selected cells
                try:
                    selected = self.missing_sheet.get_selected_cells()
                    if selected:
                        for (r, c) in selected:
                            if r < len(all_data) and c < len(all_data[r]):
                                modified_cells[(r, c)] = all_data[r][c]
                        _dbg(f"Selected cells fallback: {dict(modified_cells)}")
                except Exception as e:
                    _dbg(f"Error getting selected cells: {e}")
            
            # CTH column index
            cth_col_idx = PENDING_HEADERS.index("CTH")
            _dbg(f"CTH column index: {cth_col_idx}")
            
            # Process each modified cell
            cth_rows_to_autofill = []
            for (r, c), _old_val in modified_cells.items():
                if r >= len(all_data):
                    continue
                    
                row_data = all_data[r]
                r_dict = {h: row_data[i] if i < len(row_data) else "" for i, h in enumerate(PENDING_HEADERS)}
                
                model = r_dict.get("Model", "")
                if not model:
                    continue
                
                col_name = PENDING_HEADERS[c] if c < len(PENDING_HEADERS) else ""
                # Read the CURRENT value from sheet data (event contains the OLD value)
                new_val = str(row_data[c]).strip() if c < len(row_data) else ""
                _dbg(f"Pasted: row={r}, col={c}, col_name='{col_name}', new_value='{new_val}', old_value='{_old_val}', model='{model}'")
                
                # Auto-save the pasted value
                def save_pasted(m=model, cn=col_name, nv=new_val):
                    try:
                        from gsheets import GSheetsClient
                        client = GSheetsClient(self.cred_path.get())
                        client.update_pending_model_row(self.gsheets_url.get(), m, self.importer, {cn: nv})
                        _dbg(f"Auto-save success for {cn}={nv}")
                    except Exception as e:
                        _dbg(f"Auto-save error: {e}")
                threading.Thread(target=save_pasted, daemon=True).start()
                
                # If CTH was pasted, queue autofill
                if c == cth_col_idx and new_val:
                    cth_rows_to_autofill.append((r, model, new_val))
            
            # Trigger CTH autofill for pasted CTH values
            for r, model, cth_val in cth_rows_to_autofill:
                _dbg(f"Triggering CTH autofill for row={r}, model='{model}', CTH='{cth_val}'")
                self.root.after(100, lambda ri=r, m=model, cv=cth_val: self._do_cth_autofill(ri, m, cv))
                
        except Exception as e:
            _dbg(f"EXCEPTION: {_tb.format_exc()}")

    def _on_missing_begin_edit(self, event):
        try:
            with open("autofill_debug.log", "a", encoding="utf-8") as f:
                f.write(f"{datetime.now().isoformat()} | BEGIN_EDIT: type={type(event).__name__}, row={event.row}, col={event.column}\n")
        except Exception:
            pass
    
    def _on_missing_cell_edited(self, event):
        # Write debug log to file since try_binding swallows all exceptions silently
        import traceback as _tb
        def _dbg(msg):
            try:
                with open("autofill_debug.log", "a", encoding="utf-8") as f:
                    f.write(f"{datetime.now().isoformat()} | {msg}\n")
            except Exception:
                pass
        
        _dbg(f"EVENT FIRED. event type={type(event).__name__}, keys={list(event.keys()) if isinstance(event, dict) else 'N/A'}")
        
        try:
            # Use same direct access pattern as the working _on_cell_edited
            r = event.row
            c = event.column
            new_val = str(event.value).strip()
            
            _dbg(f"Row={r}, Col={c}, Value='{new_val}'")
            
            all_data = self.missing_sheet.get_sheet_data()
            if r >= len(all_data):
                _dbg(f"Row {r} out of bounds (len={len(all_data)})")
                return
                
            row_data = all_data[r]
            r_dict = {h: row_data[i] if i < len(row_data) else "" for i, h in enumerate(PENDING_HEADERS)}
            
            # Check permissions
            session = auth.get_current_session()
            can_add = session.get('Can_Add_Edit_Model', False) if isinstance(session, dict) else False
            can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
            
            if not can_add and not can_approve:
                _dbg("Permission denied")
                messagebox.showerror("Permission Denied", "You do not have permission to edit models.")
                self.root.after(100, self._refresh_pending_tab)
                return
            
            # Check if editable status
            status = str(r_dict.get("Status", "")).strip().lower()
            _dbg(f"Status='{status}'")
            if status and status not in ["draft", "need correction", "pending", ""]:
                messagebox.showerror("Locked", f"This row cannot be edited because its status is '{r_dict.get('Status')}'.")
                self.root.after(100, self._refresh_pending_tab)
                return
                
            model = r_dict.get("Model", "")
            if not model:
                _dbg("No model found in row")
                return
                
            col_name = PENDING_HEADERS[c]
            _dbg(f"Edited col='{col_name}' for model='{model}'")
            
            # Background auto-save
            def save_draft():
                try:
                    from gsheets import GSheetsClient
                    client = GSheetsClient(self.cred_path.get())
                    client.update_pending_model_row(self.gsheets_url.get(), model, self.importer, {col_name: new_val})
                    _dbg("Auto-save success")
                except Exception as e:
                    _dbg(f"Auto-save error: {e}")
                    
            threading.Thread(target=save_draft, daemon=True).start()
            
            # CTH Autofill
            if col_name == "CTH" and new_val:
                _dbg(f"Triggering CTH autofill for '{new_val}'")
                self.root.after(100, lambda: self._do_cth_autofill(r, model, new_val))
            else:
                _dbg(f"Not a CTH column edit (col_name='{col_name}'), skipping autofill")
        except Exception as e:
            _dbg(f"EXCEPTION: {_tb.format_exc()}")
            messagebox.showerror("Edit Error", f"Error in cell edit:\n{_tb.format_exc()}")
            
    def _do_cth_autofill(self, row_idx, model, cth_val):
        try:
            print(f"[DEBUG] _do_cth_autofill started for CTH '{cth_val}'")
            if not hasattr(self, 'importer_df') or self.importer_df is None or self.importer_df.empty:
                print(f"[DEBUG] Error: importer_df is missing or empty.")
                return
                
            df = self.importer_df
            
            # Robust CTH matching (handle .0 appended by pandas)
            clean_cth = str(cth_val).strip()
            if clean_cth.endswith('.0'):
                clean_cth = clean_cth[:-2]
            print(f"[DEBUG] Cleaned CTH target: '{clean_cth}'")
                
            df_cth = df['CTH'].astype(str).str.strip()
            df_cth = df_cth.str.replace(r'\.0$', '', regex=True)
            
            matches = df[df_cth == clean_cth]
            print(f"[DEBUG] Found {len(matches)} matches in master sheet.")
            if matches.empty:
                return
                
            duty_cols = [
                "Basic Duty Notn", "Basic Duty Notn SNo", "Basic Duty Rate", 
                "Customs Health Cess Notn", "Customs Health Cess SNo", "Customs Health Cess Rate", 
                "SWS Notification", "SWS Notification SrNo", "SWS Rate", 
                "IGST Notification", "IGST Notification SrNo", "IGST Rate", 
                "AIDC Notn (Customs)", "AIDC Notn Sr.No.(Customs)", "End Use"
            ]
            
            distinct_combos = []
            for _, row in matches.iterrows():
                combo = {}
                for col in duty_cols:
                    val = str(row.get(col, '')).strip()
                    # Normalize numeric-looking values (e.g. "12.0" -> "12")
                    if val.endswith('.0'):
                        try:
                            float(val)
                            val = val[:-2]
                        except ValueError:
                            pass
                    # Treat "0", "0.0", "0%", and empty string as equivalent ("")
                    if val in ('0', '0.0', '0%', ''):
                        val = ''
                    combo[col] = val
                if combo not in distinct_combos:
                    distinct_combos.append(combo)
                    
            print(f"[DEBUG] Reduced to {len(distinct_combos)} distinct duty combinations.")
                    
            if len(distinct_combos) == 1:
                print(f"[DEBUG] Autofilling distinct combo directly.")
                self._apply_autofill(row_idx, model, distinct_combos[0])
                messagebox.showinfo("Autofill", f"Autofilled duty fields based on CTH {cth_val}.")
            else:
                print(f"[DEBUG] Showing popup for multiple distinct combos.")
                self._show_cth_autofill_popup(row_idx, model, cth_val, distinct_combos)
        except Exception as e:
            import traceback
            print(f"[DEBUG] Exception in _do_cth_autofill:\n{traceback.format_exc()}")
            messagebox.showerror("Autofill Error", f"Error in CTH autofill:\n{traceback.format_exc()}")
            
    def _apply_autofill(self, row_idx, model, combo_dict):
        for col_name, val in combo_dict.items():
            if col_name in PENDING_HEADERS:
                c_idx = PENDING_HEADERS.index(col_name)
                self.missing_sheet.set_cell_data(row_idx, c_idx, val, redraw=False)
        self.missing_sheet.redraw()
        
        def save_autofill():
            try:
                from gsheets import GSheetsClient
                client = GSheetsClient(self.cred_path.get())
                client.update_pending_model_row(self.gsheets_url.get(), model, self.importer, combo_dict)
            except Exception as e:
                print(f"Auto-save autofill error: {e}")
        threading.Thread(target=save_autofill, daemon=True).start()
        
    def _show_cth_autofill_popup(self, row_idx, model, cth_val, combos):
        # Determine which duty fields actually differ across the candidates
        duty_cols = list(combos[0].keys())
        differing_cols = []
        for col in duty_cols:
            values_for_col = set(c.get(col, '') for c in combos)
            if len(values_for_col) > 1:
                differing_cols.append(col)
        
        # If no fields differ (shouldn't happen after dedup), autofill directly
        if not differing_cols:
            self._apply_autofill(row_idx, model, combos[0])
            messagebox.showinfo("Autofill", f"Autofilled duty fields based on CTH {cth_val}.")
            return
        
        popup = tk.Toplevel(self.root)
        popup.title(f"Multiple Duty Profiles for CTH {cth_val}")
        popup.geometry("780x600")
        popup.configure(bg="#F4F6F8")
        popup.grab_set()
        
        tk.Label(
            popup,
            text=f"CTH {cth_val} has {len(combos)} distinct duty profiles in the master sheet.",
            font=("Segoe UI", 10, "bold"), bg="#F4F6F8", fg=_DARK_TEXT
        ).pack(pady=(10, 2))
        tk.Label(
            popup,
            text=f"Showing full profiles. The {len(differing_cols)} highlighted field(s) (marked with ⚠) differ:",
            font=("Segoe UI", 9), bg="#F4F6F8", fg=_MUTED_GRAY
        ).pack(pady=(0, 10))
        
        container = tk.Frame(popup, bg=_PANEL_WHITE)
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        canvas = tk.Canvas(container, bg=_PANEL_WHITE)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=_PANEL_WHITE)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for i, combo in enumerate(combos):
            card = tk.Frame(scrollable_frame, bg="#F9FAFB", highlightbackground=_BORDER_GRAY, highlightthickness=1)
            card.pack(fill=tk.X, pady=8, padx=8)
            
            # Header row with option label and button
            header_row = tk.Frame(card, bg="#EBF5FF", height=32)
            header_row.pack(fill=tk.X)
            header_row.pack_propagate(False)
            
            tk.Label(
                header_row, text=f"Option {i+1}",
                font=("Segoe UI", 10, "bold"), bg="#EBF5FF", fg=_PRIMARY_BLUE
            ).pack(side=tk.LEFT, padx=10)
            
            btn = _brand_button(
                header_row, "✓ Use This Profile",
                lambda c=combo: [self._apply_autofill(row_idx, model, c), popup.destroy()]
            )
            btn.pack(side=tk.RIGHT, padx=10)
            
            # Fields table containing the FULL profile
            fields_frame = tk.Frame(card, bg="#F9FAFB")
            fields_frame.pack(fill=tk.X, padx=8, pady=8)
            
            for j, col in enumerate(duty_cols):
                val = combo.get(col, '')
                display_val = val if val else "(empty)"
                
                is_differing = col in differing_cols
                
                if is_differing:
                    # Highlight differing fields with light amber bg and bold text
                    row_bg = "#FEF3C7"
                    label_fg = "#B45309"
                    val_fg = "#B45309"
                    font_style = ("Segoe UI", 9, "bold")
                    display_text = f"{display_val} ⚠"
                else:
                    row_bg = "#FFFFFF" if j % 2 == 0 else "#F9FAFB"
                    label_fg = _DARK_TEXT
                    val_fg = _MUTED_GRAY if not val else _PRIMARY_BLUE
                    font_style = ("Segoe UI", 9)
                    display_text = display_val
                
                row_frame = tk.Frame(fields_frame, bg=row_bg)
                row_frame.pack(fill=tk.X)
                
                tk.Label(
                    row_frame, text=col, font=font_style,
                    bg=row_bg, fg=label_fg, width=32, anchor="w"
                ).pack(side=tk.LEFT, padx=(8, 4), pady=3)
                
                tk.Label(
                    row_frame, text=display_text, font=font_style,
                    bg=row_bg, fg=val_fg, anchor="w"
                ).pack(side=tk.LEFT, padx=(4, 8), pady=3)
            
    def _on_missing_rc_select(self, event):
        session = auth.get_current_session()
        can_approve = session.get('Can_Approve_Model', False) if isinstance(session, dict) else False
        if not can_approve:
            return
            
        r = event.row
        all_data = self.missing_sheet.get_sheet_data()
        if r >= len(all_data):
            return
            
        row_data = all_data[r]
        r_dict = {h: row_data[i] if i < len(row_data) else "" for i, h in enumerate(PENDING_HEADERS)}
        model = r_dict.get("Model", "")
        if not model:
            return
            
        popup_menu = tk.Menu(self.root, tearoff=0)
        
        def set_status(new_status):
            self.root.config(cursor="watch")
            self.root.update_idletasks()
            def update_thread():
                try:
                    from gsheets import GSheetsClient
                    from datetime import datetime
                    client = GSheetsClient(self.cred_path.get())
                    username = session.get('Username', 'User') if isinstance(session, dict) else 'User'
                    updates = {
                        "Status": new_status,
                        "Reviewed_By": username,
                        "Reviewed_At": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    client.update_pending_model_row(self.gsheets_url.get(), model, self.importer, updates)
                    self.root.after(0, lambda: [self.root.config(cursor=""), self._refresh_pending_tab()])
                except Exception as e:
                    self.root.after(0, lambda err=e: [self.root.config(cursor=""), messagebox.showerror("Error", f"Failed to update status:\n{err}")])
            threading.Thread(target=update_thread, daemon=True).start()
            
        popup_menu.add_command(label=f"Approve Model '{model}'", command=lambda: set_status("Approved"))
        popup_menu.add_command(label=f"Send back for Correction", command=lambda: set_status("Need Correction"))
        popup_menu.add_command(label=f"Reject Model", command=lambda: set_status("Rejected"))
        
        popup_menu.tk_popup(event.widget.winfo_rootx() + event.x, event.widget.winfo_rooty() + event.y)
        
    def _do_recheck_models(self):
        if not self.importer:
            return
            
        self.root.config(cursor="watch")
        if hasattr(self, 'status_label'):
            self.status_label.config(text="⚙️ Re-checking models...")
        self.root.update_idletasks()
        
        def recheck_thread():
            try:
                from gsheets import GSheetsClient
                client = GSheetsClient(self.cred_path.get())
                df, _ = client.get_sheet_data(self.gsheets_url.get(), self.importer)
                if not df.empty:
                    for col in df.columns:
                        if str(col).strip().lower() == 'model':
                            df.rename(columns={col: 'Model'}, inplace=True)
                            break
                    if 'Model' in df.columns:
                        df['Model'] = df['Model'].astype(str).str.strip()
                
                self.root.after(0, lambda d=df: self._on_recheck_success(d))
            except Exception as e:
                self.root.after(0, lambda err=e: self._on_recheck_error(err))
                
        threading.Thread(target=recheck_thread, daemon=True).start()
        
    def _on_recheck_success(self, new_df):
        self.importer_df = new_df
        
        if hasattr(self, 'session_queue') and self.session_queue:
            # Replay the entire session queue against the fresh master sheet
            self.rows_data = []
            if hasattr(self, 'sheet') and self.sheet.winfo_exists():
                self.sheet.set_sheet_data([])
            try:
                if hasattr(self, 'missing_sheet') and self.missing_sheet.winfo_exists():
                    self.missing_sheet.set_sheet_data([])
            except Exception:
                pass
            self._process_model_queue(self.session_queue)
        else:
            # Fallback if queue is empty for some reason
            self._reload_table_rows()
            
        self.root.config(cursor="")
        if hasattr(self, 'status_label') and self.status_label.winfo_exists():
            self.status_label.config(text="")
            
        messagebox.showinfo("Re-check Complete", "Successfully re-checked models against latest master.")
        
    def _on_recheck_error(self, err):
        self.root.config(cursor="")
        if hasattr(self, 'status_label') and self.status_label.winfo_exists():
            self.status_label.config(text="")
        messagebox.showerror("Error", f"Failed to reload master: {err}")
